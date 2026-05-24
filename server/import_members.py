"""
import_members.py — Import member_details.xlsm into the member_details table.

Usage:
    python server/import_members.py
"""
import os, sys
sys.path.insert(0, os.path.dirname(__file__))

import openpyxl
from datetime import datetime, date
from db import get_conn

EXCEL_PATH = os.path.join(
    os.path.dirname(__file__), '..', 'dataset', 'member_details.xlsm'
)

# Maps Excel column header → DB column name (in order, matching the sheet)
COLUMN_MAP = [
    ('SL',                                    'sl'),
    ('UID',                                   'uid'),
    ('Name',                                  'name'),
    ('Date of Initiation',                    'date_of_initiation'),
    ('BSL',                                   'bsl'),
    ('Date of Birth',                         'date_of_birth'),
    ('Date of Registration (Jigyasu)',         'date_of_registration_jigyasu'),
    ('Date of First Initiation',              'date_of_first_initiation'),
    ('Date of Second Initiation',             'date_of_second_initiation'),
    ('caste',                                 'caste'),
    ('Nationality',                           'nationality'),
    ('Qualification',                         'qualification'),
    ('Occupation',                            'occupation'),
    ('Designation',                           'designation'),
    ('Organization',                          'organization'),
    ('Profession',                            'profession'),
    ('Address Line1',                         'address_line1'),
    ('Address Line2',                         'address_line2'),
    ('Address Line3',                         'address_line3'),
    ('City',                                  'city'),
    ('Pincode',                               'pincode'),
    ('State',                                 'state'),
    ('Country',                               'country'),
    ('SN/EXT',                               'sn_ext'),
    ('ASHRAM',                               'ashram'),
    ('Email-1',                               'email1'),
    ('Email-2',                               'email2'),
    ('Mobile-1',                              'mobile1'),
    ('Mobile-2',                              'mobile2'),
    ('Landline',                              'landline'),
    ('Office Phone',                          'office_phone'),
    ('Blood Group',                           'blood_group'),
    ('Branch I. Card Received',               'branch_id_card_received'),
    ('Father Title',                          'father_title'),
    ('Father First Name',                     'father_first_name'),
    ('Father Middle Name',                    'father_middle_name'),
    ('Father Last Name',                      'father_last_name'),
    ('Father Branch',                         'father_branch'),
    ('Father BSLNO',                          'father_bslno'),
    ('Father UID',                            'father_uid'),
    ('Father DOI',                            'father_doi'),
    ('Father Phone No.',                      'father_phone'),
    ('Father City',                           'father_city'),
    ('Father State',                          'father_state'),
    ('Mother Title',                          'mother_title'),
    ('Mother First Name',                     'mother_first_name'),
    ('Mother Middle Name',                    'mother_middle_name'),
    ('Mother Last Name',                      'mother_last_name'),
    ('Mother Branch',                         'mother_branch'),
    ('Mother BSLNO',                          'mother_bslno'),
    ('Mother UID',                            'mother_uid'),
    ('Mother DOI',                            'mother_doi'),
    ('Mother Phone No.',                      'mother_phone'),
    ('Mother City',                           'mother_city'),
    ('Mother State',                          'mother_state'),
    ('Spouse Title',                          'spouse_title'),
    ('Spouse First Name',                     'spouse_first_name'),
    ('Spouse Middle Name',                    'spouse_middle_name'),
    ('Spouse Last Name',                      'spouse_last_name'),
    ('Spouse Branch',                         'spouse_branch'),
    ('Spouse BSLNO',                          'spouse_bslno'),
    ('Spouse UID',                            'spouse_uid'),
    ('Spouse DOI',                            'spouse_doi'),
    ('Spouse Phone No.',                      'spouse_phone'),
    ('Spouse City',                           'spouse_city'),
    ('Spouse State',                          'spouse_state'),
    ('Mahila Association Member',             'mahila_association_member'),
    ('Youth Member',                          'youth_member'),
    ('Associate Youth Member',                'associate_youth_member'),
    ('Junior Pre Initiate Member',            'junior_pre_initiate_member'),
    ('Senior Pre Initiate Member',            'senior_pre_initiate_member'),
    ('CRC Member',                            'crc_member'),
    ('CCA Member',                            'cca_member'),
    ('Sant-Su Member',                        'sant_su_member'),
    ('Nee (First Name)',                      'nee_first_name'),
    ('Nee (Middle Name)',                     'nee_middle_name'),
    ('Nee (Last Name)',                       'nee_last_name'),
    ('Ref-1 (Name)',                          'ref1_name'),
    ('Ref-1 (Address)',                       'ref1_address'),
    ('Ref-1 (E-mail)',                        'ref1_email'),
    ('Ref-1 (Mobile/Phone)',                  'ref1_phone'),
    ('Ref-1 (Branch Name)',                   'ref1_branch'),
    ('Ref-1 (Relation)',                      'ref1_relation'),
    ('Ref-2 (Name)',                          'ref2_name'),
    ('Ref-2 (Address)',                       'ref2_address'),
    ('Ref-3 (E-mail)',                        'ref2_email'),
    ('Ref-4 (Mobile/Phone)',                  'ref2_phone'),
    ('Ref-5 (Branch Name)',                   'ref2_branch'),
    ('Ref-6 (Relation)',                      'ref2_relation'),
    ('DOR (youth)',                           'dor_youth'),
    ('Date of Initiation (For New Initiate)', 'date_of_initiation_new'),
    ('Date of Transfer In',                   'date_transfer_in'),
    ('Transfer From (Branch)',                'transfer_from_branch'),
    ('Date of Transfer Out',                  'date_transfer_out'),
    ('Transfer To (Branch)',                  'transfer_to_branch'),
    ('Date of Expire',                        'date_of_expire'),
    ('Record Status',                         'record_status'),
    ('Profession Code',                       'profession_code'),
    ('Communication Grid Code',              'communication_grid_code'),
]

DATE_COLUMNS = {
    'date_of_initiation', 'date_of_birth', 'date_of_registration_jigyasu',
    'date_of_first_initiation', 'date_of_second_initiation',
    'father_doi', 'mother_doi', 'spouse_doi',
    'dor_youth', 'date_of_initiation_new',
    'date_transfer_in', 'date_transfer_out', 'date_of_expire',
}


def parse_date(val):
    """Return a date object or None from various possible cell formats."""
    if val is None:
        return None
    if isinstance(val, (datetime, date)):
        return val if isinstance(val, date) else val.date()
    s = str(val).strip()
    if not s:
        return None
    for fmt in ('%d-%m-%Y', '%d/%m/%Y', '%Y-%m-%d', '%m/%d/%Y', '%d-%b-%Y', '%d %b %Y'):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            pass
    return None


def clean(val, db_col):
    """Coerce a cell value to the appropriate Python type for the DB column."""
    if val is None:
        return None
    if db_col == 'sl':
        try:
            return int(val)
        except (TypeError, ValueError):
            return None
    if db_col in DATE_COLUMNS:
        return parse_date(val)
    # Everything else → string, stripped (incl. non-breaking spaces), None if empty
    s = str(val).replace('\xa0', ' ').strip() if not isinstance(val, str) else val.replace('\xa0', ' ').strip()
    return s if s else None


def build_header_index(ws, header_row=2):
    """
    Read the given header_row and return a dict {excel_header_text: col_index (0-based)}.
    Strips regular and non-breaking whitespace.
    """
    headers = {}
    for idx, cell in enumerate(next(ws.iter_rows(min_row=header_row, max_row=header_row))):
        if cell.value is not None:
            key = str(cell.value).replace('\xa0', ' ').strip()
            headers[key] = idx
    return headers


def import_members():
    path = os.path.abspath(EXCEL_PATH)
    print(f"Opening: {path}")

    wb = openpyxl.load_workbook(path, read_only=True, keep_vba=False, data_only=True)
    ws = wb.active
    print(f"Active sheet: {ws.title}")

    HEADER_ROW = 2
    DATA_START_ROW = 3
    header_index = build_header_index(ws, header_row=HEADER_ROW)

    # Resolve which Excel column index maps to each DB column
    col_mapping = []   # list of (db_col, excel_col_idx)
    for excel_hdr, db_col in COLUMN_MAP:
        if excel_hdr in header_index:
            col_mapping.append((db_col, header_index[excel_hdr]))
        else:
            print(f"  WARNING: header '{excel_hdr}' not found in sheet — column will be NULL")

    db_cols = [db_col for db_col, _ in col_mapping]
    placeholders = ', '.join(['%s'] * len(db_cols))
    col_names = ', '.join(db_cols)

    INSERT_SQL = (
        f"INSERT INTO member_details ({col_names}) VALUES ({placeholders}) "
        f"ON CONFLICT (uid) DO UPDATE SET "
        + ', '.join(f"{c} = EXCLUDED.{c}" for c in db_cols if c != 'uid')
    )

    conn = get_conn()
    cur = conn.cursor()

    inserted = 0
    skipped = 0

    for row_num, row in enumerate(ws.iter_rows(min_row=DATA_START_ROW, values_only=True), start=DATA_START_ROW):
        # Skip completely empty rows
        if all(cell is None for cell in row):
            continue

        values = []
        uid_val = None
        for db_col, col_idx in col_mapping:
            raw = row[col_idx] if col_idx < len(row) else None
            v = clean(raw, db_col)
            values.append(v)
            if db_col == 'uid':
                uid_val = v

        # Skip rows with no UID
        if not uid_val:
            skipped += 1
            continue

        try:
            cur.execute(INSERT_SQL, values)
            inserted += 1
        except Exception as e:
            print(f"  Row {row_num} (UID={uid_val}): ERROR — {e}")
            conn.rollback()
            skipped += 1
            continue

    conn.commit()
    cur.close()
    conn.close()
    wb.close()

    print(f"\nDone. Inserted/updated: {inserted}  |  Skipped: {skipped}")


if __name__ == '__main__':
    import_members()



'''
"""
import_forma.py — Import data from FormA Excel into the database.

  Sheet 'Jigyasus'   → member_details   (upsert on uid)
  Sheet 'Superhumane' → superhumane_details (upsert on uid)

Usage (from repo root):
    python server/import_forma.py

The Excel file is expected at: dataset/FormA_2026.03.28 v2_Latest.xlsm
"""
import os, sys
sys.path.insert(0, os.path.dirname(__file__))

import openpyxl
from datetime import datetime, date
from db import get_conn

EXCEL_PATH = os.path.join(
    os.path.dirname(__file__), '..', 'dataset', 'FormA_2026.03.28 v2_Latest.xlsm'
)

# ---------------------------------------------------------------------------
# Shared helpers
# ---------------------------------------------------------------------------

def parse_date(val):
    """Return a date object or None from various cell formats."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    s = str(val).replace('\xa0', ' ').strip()
    if not s or s.lower() in ('na', 'n/a', 'null', 'none'):
        return None
    for fmt in ('%d-%m-%Y', '%d/%m/%Y', '%Y-%m-%d', '%m/%d/%Y', '%d-%b-%Y', '%d %b %Y'):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            pass
    return None


def to_str(val):
    """Convert a cell value to a stripped string, or None if empty."""
    if val is None:
        return None
    if isinstance(val, (datetime, date)):
        return None  # dates should use parse_date
    s = str(val).replace('\xa0', ' ').strip()
    return s if s and s.lower() not in ('na', 'n/a', 'null', 'none') else None


def to_int(val):
    try:
        return int(val)
    except (TypeError, ValueError):
        return None


# ---------------------------------------------------------------------------
# Jigyasus → member_details
# ---------------------------------------------------------------------------

# Maps Excel header (stripped) → (db_column, type) where type is 'date'|'int'|'str'
JIGYASU_COL_MAP = [
    ('SL',                                    'sl',                          'int'),
    ('UID',                                   'uid',                         'str'),
    ('Name',                                  'name',                        'str'),
    ('Date of Initiation',                    'date_of_initiation',          'date'),
    ('BSL',                                   'bsl',                         'str'),
    ('Category',                              'category',                    'str'),
    ('Gender',                                'gender',                      'str'),
    ('Marital Status',                        'marital_status',              'str'),
    ('Previous Branch',                       'previous_branch',             'str'),
    ('Date of Birth',                         'date_of_birth',               'date'),
    ('Date of Registration (Jigyasu)',        'date_of_registration_jigyasu','date'),
    ('Date of First Initiation',             'date_of_first_initiation',    'date'),
    ('Date of Second Initiation',            'date_of_second_initiation',   'date'),
    ('caste',                                 'caste',                       'str'),
    ('Nationality',                           'nationality',                 'str'),
    ('Qualification',                         'qualification',               'str'),
    ('Occupation',                            'occupation',                  'str'),
    ('Designation',                           'designation',                 'str'),
    ('Organization',                          'organization',                'str'),
    ('Profession',                            'profession',                  'str'),
    ('Address Line1',                         'address_line1',               'str'),
    ('Address Line2',                         'address_line2',               'str'),
    ('Address Line3',                         'address_line3',               'str'),
    ('City',                                  'city',                        'str'),
    ('Pincode',                               'pincode',                     'str'),
    ('State',                                 'state',                       'str'),
    ('Country',                               'country',                     'str'),
    ('SN/EXT',                                'sn_ext',                      'str'),
    ('ASHRAM',                                'ashram',                      'str'),
    ('Email-1',                               'email1',                      'str'),
    ('Email-2',                               'email2',                      'str'),
    ('Mobile-1',                              'mobile1',                     'str'),
    ('Mobile-2',                              'mobile2',                     'str'),
    ('Landline',                              'landline',                    'str'),
    ('Office Phone',                          'office_phone',                'str'),
    ('Blood Group',                           'blood_group',                 'str'),
    ('Branch I. Card Received',               'branch_id_card_received',     'str'),
    ('Father Title',                          'father_title',                'str'),
    ('Father First Name',                     'father_first_name',           'str'),
    ('Father Middle Name',                    'father_middle_name',          'str'),
    ('Father Last Name',                      'father_last_name',            'str'),
    ('Father Branch',                         'father_branch',               'str'),
    ('Father BSLNO',                          'father_bslno',                'str'),
    ('Father UID',                            'father_uid',                  'str'),
    ('Father DOI',                            'father_doi',                  'date'),
    ('Father Phone No.',                      'father_phone',                'str'),
    ('Father City',                           'father_city',                 'str'),
    ('Father State',                          'father_state',                'str'),
    ('Mother Title',                          'mother_title',                'str'),
    ('Mother First Name',                     'mother_first_name',           'str'),
    ('Mother Middle Name',                    'mother_middle_name',          'str'),
    ('Mother Last Name',                      'mother_last_name',            'str'),
    ('Mother Branch',                         'mother_branch',               'str'),
    ('Mother BSLNO',                          'mother_bslno',                'str'),
    ('Mother UID',                            'mother_uid',                  'str'),
    ('Mother DOI',                            'mother_doi',                  'date'),
    ('Mother Phone No.',                      'mother_phone',                'str'),
    ('Mother City',                           'mother_city',                 'str'),
    ('Mother State',                          'mother_state',                'str'),
    ('Spouse Title',                          'spouse_title',                'str'),
    ('Spouse First Name',                     'spouse_first_name',           'str'),
    ('Spouse Middle Name',                    'spouse_middle_name',          'str'),
    ('Spouse Last Name',                      'spouse_last_name',            'str'),
    ('Spouse Branch',                         'spouse_branch',               'str'),
    ('Spouse BSLNO',                          'spouse_bslno',                'str'),
    ('Spouse UID',                            'spouse_uid',                  'str'),
    ('Spouse DOI',                            'spouse_doi',                  'date'),
    ('Spouse Phone No.',                      'spouse_phone',                'str'),
    ('Spouse City',                           'spouse_city',                 'str'),
    ('Spouse State',                          'spouse_state',                'str'),
    ('Mahila Association Member',             'mahila_association_member',   'str'),
    ('Youth Member',                          'youth_member',                'str'),
    ('Associate Youth Member',               'associate_youth_member',      'str'),
    ('Junior Pre Initiate Member',           'junior_pre_initiate_member',  'str'),
    ('Senior Pre Initiate Member',           'senior_pre_initiate_member',  'str'),
    ('CRC Member',                            'crc_member',                  'str'),
    ('CCA Member',                            'cca_member',                  'str'),
    ('Sant-Su Member',                        'sant_su_member',              'str'),
    ('Nee (First Name)',                      'nee_first_name',              'str'),
    ('Nee (Middle Name)',                     'nee_middle_name',             'str'),
    ('Nee (Last Name)',                       'nee_last_name',               'str'),
    ('Ref-1 (Name)',                          'ref1_name',                   'str'),
    ('Ref-1 (Address)',                       'ref1_address',                'str'),
    ('Ref-1 (E-mail)',                        'ref1_email',                  'str'),
    ('Ref-1 (Mobile/Phone)',                  'ref1_phone',                  'str'),
    ('Ref-1 (Branch Name)',                   'ref1_branch',                 'str'),
    ('Ref-1 (Relation)',                      'ref1_relation',               'str'),
    ('Ref-2 (Name)',                          'ref2_name',                   'str'),
    ('Ref-2 (Address)',                       'ref2_address',                'str'),
    ('Ref-3 (E-mail)',                        'ref2_email',                  'str'),
    ('Ref-4 (Mobile/Phone)',                  'ref2_phone',                  'str'),
    ('Ref-5 (Branch Name)',                   'ref2_branch',                 'str'),
    ('Ref-6 (Relation)',                      'ref2_relation',               'str'),
    ('DOR (youth)',                           'dor_youth',                   'date'),
    ('Date of Initiation (For New Initiate)', 'date_of_initiation_new',     'date'),
    ('Date of Transfer In',                   'date_transfer_in',            'date'),
    ('Transfer From (Branch)',               'transfer_from_branch',        'str'),
    ('Date of Transfer Out',                  'date_transfer_out',           'date'),
    ('Transfer To (Branch)',                 'transfer_to_branch',          'str'),
    ('Date of Expire',                        'date_of_expire',              'date'),
    ('Record Status',                         'record_status',               'str'),
    ('Profession Code',                       'profession_code',             'str'),
    ('Communication Grid Code',              'communication_grid_code',     'str'),
]


def import_jigyasus(wb, conn):
    ws = wb['Jigyasus']
    print("\n=== Importing Jigyasus → member_details ===")

    # Build header index from row 1
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    header_index = {}
    for col_idx, val in enumerate(header_row):
        if val is not None:
            key = str(val).replace('\xa0', ' ').strip()
            if key not in header_index:          # first occurrence wins for UID
                header_index[key] = col_idx

    # Resolve column positions
    col_mapping = []  # (db_col, col_idx, type)
    for excel_hdr, db_col, typ in JIGYASU_COL_MAP:
        if excel_hdr in header_index:
            col_mapping.append((db_col, header_index[excel_hdr], typ))
        else:
            print(f"  WARNING: header '{excel_hdr}' not found — will be NULL")

    db_cols = [db_col for db_col, _, _ in col_mapping]
    placeholders = ', '.join(['%s'] * len(db_cols))
    col_names = ', '.join(db_cols)
    insert_sql = (
        f"INSERT INTO member_details ({col_names}) VALUES ({placeholders}) "
        f"ON CONFLICT (uid) DO UPDATE SET "
        + ', '.join(f"{c} = EXCLUDED.{c}" for c in db_cols if c != 'uid')
    )

    cur = conn.cursor()
    inserted = skipped = 0

    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if all(v is None for v in row):
            continue

        values = []
        uid_val = None
        for db_col, col_idx, typ in col_mapping:
            raw = row[col_idx] if col_idx < len(row) else None
            if typ == 'date':
                v = parse_date(raw)
            elif typ == 'int':
                v = to_int(raw)
            else:
                v = to_str(raw)
            values.append(v)
            if db_col == 'uid':
                uid_val = v

        if not uid_val:
            skipped += 1
            continue

        try:
            cur.execute(insert_sql, values)
            inserted += 1
        except Exception as e:
            print(f"  Row {row_num} (UID={uid_val}): ERROR — {e}")
            conn.rollback()
            skipped += 1
            continue

    conn.commit()
    cur.close()
    print(f"  Done. Inserted/updated: {inserted}  |  Skipped: {skipped}")


# ---------------------------------------------------------------------------
# Superhumane → superhumane_details
# The sheet has merged/repeated header labels so we use fixed column indices.
# Header is on row 4; data starts row 5.
# ---------------------------------------------------------------------------

# (db_col, col_index_0based, type)
SUPERHUMANE_COL_MAP = [
    ('sno',                 0,  'int'),
    ('member_type',         1,  'str'),
    ('name',                2,  'str'),
    ('form_check',          3,  'str'),
    ('uid_check',           4,  'str'),
    ('bsl',                 5,  'str'),
    ('gender',              6,  'str'),
    ('comments',            7,  'str'),
    ('uid',                 8,  'str'),
    ('date_of_birth',       9,  'date'),
    # col 10 = Age (formula, skip)
    ('phase',               11, 'int'),
    ('branch',              12, 'str'),
    ('father_name',         13, 'str'),
    ('father_contact',      14, 'str'),
    ('father_uid',          15, 'str'),
    ('father_doi',          16, 'date'),
    ('mother_name',         17, 'str'),
    ('mother_contact',      18, 'str'),
    ('mother_uid',          19, 'str'),
    ('mother_doi',          20, 'date'),
    ('address',             21, 'str'),
    ('grandfather_name',    22, 'str'),
    ('grandfather_uid',     23, 'str'),
    ('grandfather_contact', 24, 'str'),
    ('grandmother_name',    25, 'str'),
    ('grandmother_uid',     26, 'str'),
    ('grandmother_contact', 27, 'str'),
    ('date_entry_scheme',   28, 'date'),
    ('date_exit_scheme',    29, 'date'),
]


def import_superhumane(wb, conn):
    ws = wb['Superhumane']
    print("\n=== Importing Superhumane → superhumane_details ===")

    db_cols = [db_col for db_col, _, _ in SUPERHUMANE_COL_MAP]
    placeholders = ', '.join(['%s'] * len(db_cols))
    col_names = ', '.join(db_cols)
    insert_sql = (
        f"INSERT INTO superhumane_details ({col_names}) VALUES ({placeholders}) "
        f"ON CONFLICT (uid) DO UPDATE SET "
        + ', '.join(f"{c} = EXCLUDED.{c}" for c in db_cols if c != 'uid')
    )

    cur = conn.cursor()
    inserted = skipped = 0

    for row_num, row in enumerate(ws.iter_rows(min_row=5, values_only=True), start=5):
        if all(v is None for v in row):
            continue

        values = []
        uid_val = None
        for db_col, col_idx, typ in SUPERHUMANE_COL_MAP:
            raw = row[col_idx] if col_idx < len(row) else None
            # Skip formula strings (Age column formula spills)
            if isinstance(raw, str) and raw.startswith('='):
                raw = None
            if typ == 'date':
                v = parse_date(raw)
            elif typ == 'int':
                v = to_int(raw)
            else:
                v = to_str(raw)
            values.append(v)
            if db_col == 'uid':
                uid_val = v

        if not uid_val:
            skipped += 1
            continue

        try:
            cur.execute(insert_sql, values)
            inserted += 1
        except Exception as e:
            print(f"  Row {row_num} (UID={uid_val}): ERROR — {e}")
            conn.rollback()
            skipped += 1
            continue

    conn.commit()
    cur.close()
    print(f"  Done. Inserted/updated: {inserted}  |  Skipped: {skipped}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    path = os.path.abspath(EXCEL_PATH)
    print(f"Opening: {path}")

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    print(f"Sheets found: {wb.sheetnames}")

    conn = get_conn()
    try:
        import_jigyasus(wb, conn)
        import_superhumane(wb, conn)
    finally:
        conn.close()
        wb.close()

    print("\nAll done.")


if __name__ == '__main__':
    main()

'''
