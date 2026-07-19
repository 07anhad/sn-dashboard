"""
app.py — Flask server: serves static frontend + REST API
"""
import os, sys, logging, shutil, smtplib
import datetime as _dt
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from logging.handlers import TimedRotatingFileHandler
sys.path.insert(0, os.path.dirname(__file__))


class _WinSafeRotatingHandler(TimedRotatingFileHandler):
    """TimedRotatingFileHandler that works on Windows by copy+truncate instead of rename."""
    def rotate(self, source, dest):
        try:
            shutil.copy2(source, dest)
            with open(source, 'w', encoding='utf-8'):
                pass  # truncate original
        except Exception:
            pass  # never crash the server over a log rotation failure

from flask import Flask, send_from_directory, jsonify, request
from flask_cors import CORS
from db import query, execute

app = Flask(__name__, static_folder=None)
CORS(app)
app.config['MAX_CONTENT_LENGTH'] = 100 * 1024 * 1024  # 100 MB upload limit
app.secret_key = os.environ.get('SECRET_KEY', 'dev-only-insecure-key')

ROOT = os.path.join(os.path.dirname(__file__), '..')

# ═══════════════════════════════════════════
# Auto-initialise DB schema on startup
# ═══════════════════════════════════════════
def _init_db():
    schema_path = os.path.join(os.path.dirname(__file__), 'schema.sql')
    try:
        with open(schema_path, 'r', encoding='utf-8') as f:
            sql = f.read()
        from db import get_conn
        conn = get_conn()
        with conn.cursor() as cur:
            cur.execute(sql)
        conn.commit()
        conn.close()
        print('[startup] DB schema initialised OK')
    except Exception as e:
        print(f'[startup] DB init error: {e}')

_init_db()

# ═══════════════════════════════════════════
# Email — Gmail SMTP
# Set GMAIL_APP_PASSWORD env var (Gmail App Password for soaminagarbranch@gmail.com)
# ═══════════════════════════════════════════
GMAIL_USER     = 'soaminagarbranch@gmail.com'
GMAIL_PASSWORD = os.environ.get('GMAIL_APP_PASSWORD', '')

def send_email(to_addr, subject, html_body):
    """Send an HTML email via Gmail SMTP. Logs on failure, never crashes server."""
    if not to_addr:
        return
    if not GMAIL_PASSWORD:
        print(f"[EMAIL DEV] Would send to {to_addr}: {subject}", flush=True)
        return
    try:
        msg = MIMEMultipart('alternative')
        msg['Subject'] = subject
        msg['From']    = f'Soaminagar Branch Delhi <{GMAIL_USER}>'
        msg['To']      = to_addr
        msg.attach(MIMEText(html_body, 'html', 'utf-8'))
        with smtplib.SMTP('smtp.gmail.com', 587, timeout=10) as s:
            s.starttls()
            s.login(GMAIL_USER, GMAIL_PASSWORD)
            s.sendmail(GMAIL_USER, [to_addr], msg.as_string())
        logging.getLogger('audit').info(f"EMAIL_SENT to={to_addr} subject={subject}")
        print(f"[EMAIL] Sent to {to_addr}", flush=True)
    except Exception as e:
        logging.getLogger('audit').warning(f"EMAIL_FAILED to={to_addr} err={e}")
        print(f"[EMAIL ERROR] {e}", flush=True)

@app.route('/api/test-email')
def test_email():
    """Dev-only: send a test email."""
    send_email(GMAIL_USER, 'Test Email — Satsang Portal', '<h2>It works!</h2><p>Gmail SMTP is configured correctly.</p>')
    return jsonify({'ok': True, 'gmail_password_set': bool(GMAIL_PASSWORD)})

# ═══════════════════════════════════════════
# Audit logger — writes to logs/audit_YYYY-MM-DD.log
# ═══════════════════════════════════════════
_LOG_DIR = os.path.join(ROOT, 'logs')
os.makedirs(_LOG_DIR, exist_ok=True)

_audit_logger = logging.getLogger('audit')
_audit_logger.setLevel(logging.INFO)
_audit_logger.propagate = False
_handler = _WinSafeRotatingHandler(
    os.path.join(_LOG_DIR, 'audit.log'),
    when='midnight', backupCount=90, encoding='utf-8'
)
_handler.suffix = '%Y-%m-%d'
_handler.setFormatter(logging.Formatter('%(asctime)s | %(message)s', datefmt='%Y-%m-%d %H:%M:%S'))
_audit_logger.addHandler(_handler)

def audit(action: str, detail: str = ''):
    """Log a user action. Reads actor info from the X-User header sent by the frontend."""
    actor = request.headers.get('X-User', 'unknown')
    ip = request.headers.get('X-Forwarded-For', request.remote_addr or 'unknown')
    msg = f"actor={actor} | ip={ip} | action={action}"
    if detail:
        msg += f" | detail={detail}"
    _audit_logger.info(msg)

# ═══════════════════════════════════════════
# BSL (Branch Serial Number) Auto-Assignment
# Pattern: Initiated=1xxx/2xxx, Jigyasu=3xxx, Superhumane=4xxx
# ═══════════════════════════════════════════
def determine_member_category(row):
    """Determine member category based on registration data."""
    # First check if member_type was explicitly selected in the form
    if row.get('member_type'):
        return row['member_type']
    # Fallback: infer from dates if member_type not set
    if row.get('date_of_initiation'):
        return 'Initiated'
    if row.get('date_of_registration_jigyasu'):
        return 'Jigyasu'
    # Default to Jigyasu for new registrations without initiation
    return 'Jigyasu'

def get_next_bsl(category):
    """
    Generate the next BSL number based on member category.
    BSL Pattern:
      - Initiated: starts with 1 or 2 (we use the next available in 1000-2999 range)
      - Jigyasu: starts with 3 (3000-3999 range)
      - Superhumane: starts with 4 (4000-4999 range)
    """
    if category == 'Initiated':
        # Initiated members use 1xxx and 2xxx range
        # Find max BSL starting with 1 or 2 from member_details
        result = query("""
            SELECT MAX(CAST(bsl AS INTEGER)) as max_bsl 
            FROM member_details 
            WHERE bsl ~ '^[12][0-9]{3}$'
        """, one=True)
        max_bsl = result['max_bsl'] if result and result['max_bsl'] else 999
        return str(max_bsl + 1)
    
    elif category == 'Jigyasu':
        # Jigyasu members use 3xxx range
        result = query("""
            SELECT MAX(CAST(bsl AS INTEGER)) as max_bsl 
            FROM member_details 
            WHERE bsl ~ '^3[0-9]{3}$'
        """, one=True)
        max_bsl = result['max_bsl'] if result and result['max_bsl'] else 2999
        return str(max_bsl + 1)
    
    elif category == 'Superhumane':
        # Superhumane use 4xxx range
        result = query("""
            SELECT MAX(CAST(bsl AS INTEGER)) as max_bsl 
            FROM superhumane_details 
            WHERE bsl ~ '^4[0-9]{3}$'
        """, one=True)
        max_bsl = result['max_bsl'] if result and result['max_bsl'] else 3999
        return str(max_bsl + 1)
    
    # Fallback: return a high number in 9xxx range
    return '9001'

# ═══════════════════════════════════════════
# Static file serving
# ═══════════════════════════════════════════
@app.route('/')
@app.route('/index.html')
def index():
    return send_from_directory(os.path.join(ROOT, 'html'), 'index.html')

@app.route('/dashboard.html')
def dashboard():
    return send_from_directory(os.path.join(ROOT, 'html'), 'dashboard.html')

@app.route('/html/<path:filename>')
def html_files(filename):
    return send_from_directory(os.path.join(ROOT, 'html'), filename)

@app.route('/css/<path:filename>')
def css_files(filename):
    return send_from_directory(os.path.join(ROOT, 'css'), filename)

@app.route('/js/<path:filename>')
def js_files(filename):
    return send_from_directory(os.path.join(ROOT, 'js'), filename)

@app.route('/dataset/<path:filename>')
def dataset_files(filename):
    return send_from_directory(os.path.join(ROOT, 'dataset'), filename)

@app.route('/manifest.json')
def manifest():
    return send_from_directory(ROOT, 'manifest.json', mimetype='application/manifest+json')

@app.route('/sw.js')
def service_worker():
    resp = send_from_directory(ROOT, 'sw.js', mimetype='application/javascript')
    resp.headers['Service-Worker-Allowed'] = '/'
    resp.headers['Cache-Control'] = 'no-cache'
    return resp

@app.route('/icons/<path:filename>')
def icon_files(filename):
    return send_from_directory(os.path.join(ROOT, 'icons'), filename)

# ═══════════════════════════════════════════
# AUTH API
# ═══════════════════════════════════════════
@app.route('/api/debug/email-lookup')
def debug_email_lookup():
    email = (request.args.get('email') or '').strip().lower()
    if not email:
        return jsonify({'error': 'Pass ?email=...'})
    exact = query("SELECT uid, email1, email2 FROM member_details WHERE email1=%s OR email2=%s", (email, email))
    ilike = query("SELECT uid, email1, email2 FROM member_details WHERE LOWER(TRIM(email1))=%s OR LOWER(TRIM(email2))=%s", (email, email))
    return jsonify({'exact_match': exact, 'ilike_match': ilike})
@app.route('/api/auth/login', methods=['POST'])
def login():
    data = request.json or {}
    username = (data.get('username') or '').strip()
    password = data.get('password') or ''
    role = (data.get('role') or '').strip()

    # 'admin' toggle on login page covers both admin and superadmin
    if role == 'admin':
        user = query(
            "SELECT id,username,name,role,email,member_id FROM users WHERE (username=%s OR email=%s) AND password=%s AND role IN ('admin','superadmin')",
            (username, username, password),
            one=True
        )
    else:
        user = query(
            "SELECT id,username,name,role,email,member_id FROM users WHERE (username=%s OR email=%s) AND password=%s AND role=%s",
            (username, username, password, role),
            one=True
        )
    if user:
        user_dict = dict(user)
        if user_dict['role'] == 'member' and user_dict.get('email'):
            email_lower = user_dict['email'].strip().lower()
            # Prefer email1 match first, only use email2 as fallback
            md = query(
                "SELECT uid FROM member_details WHERE LOWER(TRIM(email1))=%s",
                (email_lower,), one=True
            )
            if not md:
                md = query(
                    "SELECT uid FROM member_details WHERE LOWER(TRIM(email2))=%s",
                    (email_lower,), one=True
                )
            user_dict['member_uid'] = md['uid'] if md else None
        else:
            user_dict['member_uid'] = None
        audit('LOGIN_SUCCESS', f"username={username} role={user_dict['role']} member_uid={user_dict.get('member_uid')}")
        return jsonify({'ok': True, 'user': user_dict})
    audit('LOGIN_FAILED', f"username={username} attempted_role={role}")
    return jsonify({'ok': False, 'error': 'Invalid credentials'}), 401

# ── OTP email helper (uses shared send_email / Gmail) ────────────────────────
def _send_otp_email(to_email, code):
    if not GMAIL_PASSWORD:
        # Dev mode: no App Password set — print code to console for local testing
        print(f"[OTP DEV] Code for {to_email}: {code}", flush=True)
        return
    send_email(
        to_addr   = to_email,
        subject   = f'{code} — Your Soaminagar Branch login code',
        html_body = f"""
    <div style="font-family:sans-serif;max-width:480px;margin:auto;padding:32px;background:#f4f6fb;border-radius:12px">
      <h2 style="color:#e07b29;margin:0 0 8px">Soaminagar Branch Delhi</h2>
      <p style="color:#444;margin:0 0 24px">Your one-time login code:</p>
      <div style="font-size:2.5rem;font-weight:700;letter-spacing:12px;color:#1c1f2e;text-align:center;
                  background:#fff;border-radius:8px;padding:20px 0;margin-bottom:24px">{code}</div>
      <p style="color:#777;font-size:0.85rem;margin:0">This code expires in 10 minutes.<br>
      If you did not request this, please ignore this email.</p>
    </div>"""
    )

@app.route('/api/auth/send-otp', methods=['POST'])
def send_otp():
    import random
    data     = request.json or {}
    username = (data.get('username') or '').strip()
    password = data.get('password') or ''
    role     = (data.get('role') or '').strip()

    # Validate credentials first (same logic as login)
    if role == 'admin':
        user = query(
            "SELECT id,username,name,role,email,member_id FROM users WHERE (username=%s OR email=%s) AND password=%s AND role IN ('admin','superadmin')",
            (username, username, password), one=True
        )
    else:
        user = query(
            "SELECT id,username,name,role,email,member_id FROM users WHERE (username=%s OR email=%s) AND password=%s AND role=%s",
            (username, username, password, role), one=True
        )

    if not user:
        return jsonify({'ok': False, 'error': 'Invalid credentials.'}), 401

    email = (user.get('email') or '').strip()
    if not email:
        return jsonify({'ok': False, 'error': 'No email address linked to this account. Contact an administrator.'}), 400

    # Invalidate any existing unused codes for this email
    execute("UPDATE otp_tokens SET used=TRUE WHERE email=%s AND used=FALSE", (email,))

    code = str(random.randint(100000, 999999))
    execute(
        "INSERT INTO otp_tokens (email, code, expires_at) VALUES (%s, %s, NOW() + INTERVAL '10 minutes')",
        (email, code)
    )

    _send_otp_email(email, code)

    # Return masked email so frontend can display it
    parts  = email.split('@')
    masked = parts[0][:2] + '***@' + parts[1]
    audit('OTP_SENT', f"username={username} email={masked}")
    return jsonify({'ok': True, 'maskedEmail': masked})

@app.route('/api/auth/verify-otp', methods=['POST'])
def verify_otp():
    data     = request.json or {}
    username = (data.get('username') or '').strip()
    password = data.get('password') or ''
    role     = (data.get('role') or '').strip()
    code     = (data.get('code') or '').strip()

    # Re-validate credentials
    if role == 'admin':
        user = query(
            "SELECT id,username,name,role,email,member_id FROM users WHERE (username=%s OR email=%s) AND password=%s AND role IN ('admin','superadmin')",
            (username, username, password), one=True
        )
    else:
        user = query(
            "SELECT id,username,name,role,email,member_id FROM users WHERE (username=%s OR email=%s) AND password=%s AND role=%s",
            (username, username, password, role), one=True
        )

    if not user:
        return jsonify({'ok': False, 'error': 'Invalid credentials.'}), 401

    email = (user.get('email') or '').strip()
    # Debug: log what we're checking
    logging.warning(f"[OTP DEBUG] email={email!r} code={code!r}")
    latest = query(
        "SELECT id, code, used, expires_at FROM otp_tokens WHERE email=%s ORDER BY created_at DESC LIMIT 1",
        (email,), one=True
    )
    logging.warning(f"[OTP DEBUG] latest token={latest}")
    token = query(
        "SELECT id FROM otp_tokens WHERE email=%s AND code=%s AND used=FALSE AND expires_at > NOW() ORDER BY created_at DESC LIMIT 1",
        (email, code), one=True
    )
    if not token:
        return jsonify({'ok': False, 'error': 'Invalid or expired code. Please try again.'}), 401

    # Mark token used
    execute("UPDATE otp_tokens SET used=TRUE WHERE id=%s", (token['id'],))

    user_dict = dict(user)
    if user_dict['role'] == 'member' and email:
        email_lower = email.lower()
        md = query("SELECT uid FROM member_details WHERE LOWER(TRIM(email1))=%s", (email_lower,), one=True)
        if not md:
            md = query("SELECT uid FROM member_details WHERE LOWER(TRIM(email2))=%s", (email_lower,), one=True)
        user_dict['member_uid'] = md['uid'] if md else None
    else:
        user_dict['member_uid'] = None

    audit('OTP_LOGIN_SUCCESS', f"username={username} role={user_dict['role']}")
    return jsonify({'ok': True, 'user': user_dict})

@app.route('/api/auth/signup', methods=['POST'])
def signup():
    data = request.json or {}
    name = (data.get('name') or '').strip()
    username = (data.get('username') or '').strip().lower()
    email = (data.get('email') or '').strip().lower()
    password = data.get('password') or ''

    if not name or not username or not email or len(password) < 6:
        return jsonify({'ok': False, 'error': 'All fields required. Password min 6 chars.'}), 400

    # Check duplicates
    existing = query("SELECT id FROM users WHERE username=%s OR email=%s", (username, email))
    if existing:
        return jsonify({'ok': False, 'error': 'Username or email already taken.'}), 409

    # Generate member ID
    count = query("SELECT count(*) as c FROM users WHERE role='member'", one=True)
    member_id = 'M-' + str(10000 + (count['c'] if count else 0) + 1).zfill(5)

    user = execute(
        "INSERT INTO users (username,password,role,name,email,member_id) VALUES (%s,%s,'member',%s,%s,%s) RETURNING id,username,name,role,email,member_id",
        (username, password, name, email, member_id),
        returning=True
    )
    audit('SIGNUP', f"new_user={username} email={email}")
    return jsonify({'ok': True, 'user': user}), 201

@app.route('/api/auth/reset-password', methods=['POST'])
def reset_password():
    data = request.json or {}
    username    = (data.get('username') or '').strip().lower()
    new_password = data.get('newPassword') or ''
    code        = (data.get('code') or '').strip()

    if len(new_password) < 6:
        return jsonify({'ok': False, 'error': 'Password min 6 chars.'}), 400

    user = query("SELECT id, email FROM users WHERE username=%s", (username,), one=True)
    if not user:
        return jsonify({'ok': False, 'error': 'Username not found.'}), 404

    # If code provided, verify OTP
    if code:
        email = (user.get('email') or '').strip()
        if not email:
            return jsonify({'ok': False, 'error': 'No email linked to this account.'}), 400
        token = query(
            "SELECT id FROM otp_tokens WHERE email=%s AND code=%s AND used=FALSE AND expires_at > NOW() ORDER BY created_at DESC LIMIT 1",
            (email, code), one=True
        )
        if not token:
            return jsonify({'ok': False, 'error': 'Invalid or expired code.'}), 401
        execute("UPDATE otp_tokens SET used=TRUE WHERE id=%s", (token['id'],))

    execute("UPDATE users SET password=%s WHERE username=%s", (new_password, username))
    audit('RESET_PASSWORD', f"target_user={username}")
    return jsonify({'ok': True})

@app.route('/api/auth/forgot-send-otp', methods=['POST'])
def forgot_send_otp():
    import random
    data     = request.json or {}
    username = (data.get('username') or '').strip().lower()

    user = query("SELECT id, email FROM users WHERE username=%s", (username,), one=True)
    if not user:
        return jsonify({'ok': False, 'error': 'Username not found.'}), 404

    email = (user.get('email') or '').strip()
    if not email:
        return jsonify({'ok': False, 'error': 'No email address linked to this account. Contact an administrator.'}), 400

    execute("UPDATE otp_tokens SET used=TRUE WHERE email=%s AND used=FALSE", (email,))
    code = str(random.randint(100000, 999999))
    execute(
        "INSERT INTO otp_tokens (email, code, expires_at) VALUES (%s, %s, NOW() + INTERVAL '10 minutes')",
        (email, code)
    )

    _send_otp_email(email, code)

    parts  = email.split('@')
    masked = parts[0][:2] + '***@' + parts[1]
    audit('FORGOT_OTP_SENT', f"username={username} email={masked}")
    return jsonify({'ok': True, 'maskedEmail': masked})

# ═══════════════════════════════════════════
# ZONES API
# ═══════════════════════════════════════════
@app.route('/api/zones')
def get_zones():
    return jsonify(query("SELECT * FROM zones ORDER BY id"))

@app.route('/api/zones', methods=['POST'])
def add_zone():
    d = request.json
    row = execute(
        "INSERT INTO zones (name,code,active,member_count,incharge,phone) VALUES (%s,%s,%s,%s,%s,%s) RETURNING *",
        (d['name'], d['code'], d.get('active', True), d.get('member_count', 0), d.get('incharge',''), d.get('phone','')),
        returning=True
    )
    audit('ADD_ZONE', f"name={d['name']} code={d['code']}")
    return jsonify(row), 201

@app.route('/api/zones/<int:id>', methods=['PUT'])
def update_zone(id):
    d = request.json
    execute(
        "UPDATE zones SET name=%s,code=%s,active=%s,member_count=%s,incharge=%s,phone=%s WHERE id=%s",
        (d['name'], d['code'], d.get('active', True), d.get('member_count', 0), d.get('incharge',''), d.get('phone',''), id)
    )
    audit('EDIT_ZONE', f"id={id} name={d['name']}")
    return jsonify({'ok': True})

@app.route('/api/zones/<int:id>', methods=['DELETE'])
def delete_zone(id):
    execute("DELETE FROM zones WHERE id=%s", (id,))
    audit('DELETE_ZONE', f"id={id}")
    return jsonify({'ok': True})

# ═══════════════════════════════════════════
# BRANCHES API
# ═══════════════════════════════════════════
# ═══════════════════════════════════════════
# MEMBERS API
# ═══════════════════════════════════════════
@app.route('/api/members')
def get_members():
    search = (request.args.get('q') or '').strip()
    page   = max(1, int(request.args.get('page', 1)))
    limit  = min(10000, int(request.args.get('limit', 5000)))
    offset = (page - 1) * limit

    if search:
        like = f'%{search}%'
        rows = query(
            """SELECT * FROM member_details
               WHERE name ILIKE %s OR uid ILIKE %s OR mobile1 ILIKE %s OR bsl ILIKE %s
               ORDER BY name LIMIT %s OFFSET %s""",
            (like, like, like, like, limit, offset)
        )
        total = query(
            """SELECT count(*) as c FROM member_details
               WHERE name ILIKE %s OR uid ILIKE %s OR mobile1 ILIKE %s OR bsl ILIKE %s""",
            (like, like, like, like), one=True
        )['c']
    else:
        rows  = query("SELECT * FROM member_details ORDER BY name LIMIT %s OFFSET %s", (limit, offset))
        total = query("SELECT count(*) as c FROM member_details", one=True)['c']

    return jsonify({'members': rows, 'total': total, 'page': page, 'limit': limit})

@app.route('/api/member-types')
def get_member_types():
    rows = query("SELECT name FROM member_types ORDER BY id")
    return jsonify([r['name'] for r in rows])

@app.route('/api/members', methods=['POST'])
def add_member():
    import psycopg2
    d = request.json
    uid = (d.get('uid') or '').strip()
    if not uid:
        return jsonify({'ok': False, 'error': 'UID is required.'}), 400
    try:
        execute(
            """INSERT INTO member_details
               (uid, bsl, name, mobile1, email1, city, state, record_status)
               VALUES (%s, %s, %s, %s, %s, %s, %s, %s)""",
            (uid, d.get('bslno',''), d.get('name',''), d.get('mobile',''),
             d.get('email',''), d.get('city',''), d.get('state',''),
             d.get('status','Activated'))
        )
        audit('ADD_MEMBER', f"uid={uid} name={d.get('name','')}")
        return jsonify({'ok': True, 'uid': uid}), 201
    except psycopg2.errors.UniqueViolation:
        return jsonify({'ok': False, 'error': f"A member with UID '{uid}' already exists."}), 409
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500

@app.route('/api/members/<uid>', methods=['PUT'])
@app.route('/api/members/<uid>/self', methods=['PUT'])
def update_member(uid):
    import re
    d = request.json or {}
    is_self_edit = request.path.endswith('/self')

    if is_self_edit:
        # Member self-edit: only allowed contact/address/professional fields
        allowed = {
            'mobile1':        d.get('mobile'),
            'mobile2':        d.get('mobile2'),
            'landline':       d.get('landline'),
            'office_phone':   d.get('officePhone'),
            'email1':         d.get('email'),
            'email2':         d.get('email2'),
            'address_line1':  d.get('addressLine1'),
            'address_line2':  d.get('addressLine2'),
            'address_line3':  d.get('addressLine3'),
            'city':           d.get('city'),
            'pincode':        d.get('pincode'),
            'state':          d.get('state'),
            'country':        d.get('country'),
            'qualification':  d.get('qualification'),
            'occupation':     d.get('occupation'),
            'designation':    d.get('designation'),
            'organization':   d.get('organization'),
            'profession':     d.get('profession'),
        }
        set_clause = ', '.join(f"{col}=%s" for col in allowed)
        values = list(allowed.values()) + [uid]
        execute(f"UPDATE member_details SET {set_clause} WHERE uid=%s", values)
        audit('SELF_EDIT_MEMBER', f"uid={uid}")
        return jsonify({'ok': True})

    # Admin full update below
    d = request.json

    # Status-only update (toggle activate/deactivate)
    if set(d.keys()) <= {'status'}:
        execute("UPDATE member_details SET record_status=%s WHERE uid=%s",
                (d.get('status', 'Activated'), uid))
        audit('UPDATE_MEMBER_STATUS', f"uid={uid} status={d.get('status','Activated')}")
        return jsonify({'ok': True})

    # Full field update
    or_none = lambda k: d.get(k) or None
    execute("""
        UPDATE member_details SET
          name=%s, bsl=%s,
          date_of_initiation=%s, date_of_birth=%s,
          date_of_registration_jigyasu=%s,
          date_of_first_initiation=%s, date_of_second_initiation=%s,
          blood_group=%s, caste=%s, nationality=%s,
          ashram=%s, sn_ext=%s, branch_id_card_received=%s,
          record_status=%s,
          mobile1=%s, mobile2=%s, landline=%s, office_phone=%s,
          email1=%s, email2=%s,
          address_line1=%s, address_line2=%s, address_line3=%s,
          city=%s, pincode=%s, state=%s, country=%s,
          qualification=%s, occupation=%s, designation=%s,
          organization=%s, profession=%s,
          profession_code=%s, communication_grid_code=%s,
          mahila_association_member=%s, youth_member=%s, associate_youth_member=%s,
          junior_pre_initiate_member=%s, senior_pre_initiate_member=%s,
          crc_member=%s, cca_member=%s, sant_su_member=%s,
          nee_first_name=%s, nee_middle_name=%s, nee_last_name=%s,
          father_title=%s, father_first_name=%s, father_middle_name=%s, father_last_name=%s,
          father_branch=%s, father_bslno=%s, father_uid=%s, father_doi=%s,
          father_phone=%s, father_city=%s, father_state=%s,
          mother_title=%s, mother_first_name=%s, mother_middle_name=%s, mother_last_name=%s,
          mother_branch=%s, mother_bslno=%s, mother_uid=%s, mother_doi=%s,
          mother_phone=%s, mother_city=%s, mother_state=%s,
          spouse_title=%s, spouse_first_name=%s, spouse_middle_name=%s, spouse_last_name=%s,
          spouse_branch=%s, spouse_bslno=%s, spouse_uid=%s, spouse_doi=%s,
          spouse_phone=%s, spouse_city=%s, spouse_state=%s,
          ref1_name=%s, ref1_address=%s, ref1_email=%s, ref1_phone=%s,
          ref1_branch=%s, ref1_relation=%s,
          ref2_name=%s, ref2_address=%s, ref2_email=%s, ref2_phone=%s,
          ref2_branch=%s, ref2_relation=%s,
          dor_youth=%s, date_of_initiation_new=%s,
          date_transfer_in=%s, transfer_from_branch=%s,
          date_transfer_out=%s, transfer_to_branch=%s,
          date_of_expire=%s
        WHERE uid=%s
    """, (
        d.get('name'),              d.get('bslno'),
        or_none('dateOfInitiation'),or_none('dateOfBirth'),
        or_none('dateOfRegistration'),
        or_none('dateOfFirstInitiation'), or_none('dateOfSecondInitiation'),
        d.get('bloodGroup'),        d.get('caste'),         d.get('nationality'),
        d.get('ashram'),            d.get('snExt'),         d.get('branchIdCard'),
        d.get('status', 'Activated'),
        d.get('mobile'),            d.get('mobile2'),       d.get('landline'),      d.get('officePhone'),
        d.get('email'),             d.get('email2'),
        d.get('addressLine1'),      d.get('addressLine2'),  d.get('addressLine3'),
        d.get('city'),              d.get('pincode'),       d.get('state'),         d.get('country'),
        d.get('qualification'),     d.get('occupation'),    d.get('designation'),
        d.get('organization'),      d.get('profession'),
        d.get('professionCode'),    d.get('commGridCode'),
        d.get('mahila'),            d.get('youth'),         d.get('assocYouth'),
        d.get('jrPreInit'),         d.get('srPreInit'),
        d.get('crc'),               d.get('cca'),           d.get('santSu'),
        d.get('neeFirst'),          d.get('neeMiddle'),     d.get('neeLast'),
        d.get('fatherTitle'),       d.get('fatherFirstName'), d.get('fatherMiddleName'), d.get('fatherLastName'),
        d.get('fatherBranch'),      d.get('fatherBslno'),   d.get('fatherUid'),     or_none('fatherDoi'),
        d.get('fatherPhone'),       d.get('fatherCity'),    d.get('fatherState'),
        d.get('motherTitle'),       d.get('motherFirstName'), d.get('motherMiddleName'), d.get('motherLastName'),
        d.get('motherBranch'),      d.get('motherBslno'),   d.get('motherUid'),     or_none('motherDoi'),
        d.get('motherPhone'),       d.get('motherCity'),    d.get('motherState'),
        d.get('spouseTitle'),       d.get('spouseFirstName'), d.get('spouseMiddleName'), d.get('spouseLastName'),
        d.get('spouseBranch'),      d.get('spouseBslno'),   d.get('spouseUid'),     or_none('spouseDoi'),
        d.get('spousePhone'),       d.get('spouseCity'),    d.get('spouseState'),
        d.get('ref1Name'),          d.get('ref1Address'),   d.get('ref1Email'),     d.get('ref1Phone'),
        d.get('ref1Branch'),        d.get('ref1Relation'),
        d.get('ref2Name'),          d.get('ref2Address'),   d.get('ref2Email'),     d.get('ref2Phone'),
        d.get('ref2Branch'),        d.get('ref2Relation'),
        or_none('dorYouth'),        or_none('dateOfInitiationNew'),
        or_none('dateTransferIn'),  d.get('transferFromBranch'),
        or_none('dateTransferOut'), d.get('transferToBranch'),
        or_none('dateOfExpire'),
        uid
    ))
    audit('EDIT_MEMBER', f"uid={uid} name={d.get('name','')}")
    return jsonify({'ok': True})

@app.route('/api/members/<uid>', methods=['DELETE'])
def delete_member(uid):
    execute("DELETE FROM members WHERE uid=%s", (uid,))
    audit('DELETE_MEMBER', f"uid={uid}")
    return jsonify({'ok': True})

# ═══════════════════════════════════════════
# REG LINKS API
# ═══════════════════════════════════════════
@app.route('/api/reg-links')
def get_reg_links():
    return jsonify(query("SELECT * FROM reg_links ORDER BY id"))

@app.route('/api/reg-links', methods=['POST'])
def add_reg_link():
    d = request.json
    row = execute(
        "INSERT INTO reg_links (title,code,url,active,max_uses,used_count,expiry,created_on) VALUES (%s,%s,%s,%s,%s,%s,%s,%s) RETURNING *",
        (d['title'], d['code'], d.get('url',''), d.get('active',True), d.get('maxUses',0), d.get('usedCount',0), d.get('expiry'), d.get('createdOn')),
        returning=True
    )
    audit('ADD_REG_LINK', f"title={d['title']} code={d['code']}")
    return jsonify(row), 201

# AFTER — convert empty string to None so Postgres gets NULL instead of ''
@app.route('/api/reg-links/<int:id>', methods=['PUT'])
def update_reg_link(id):
    d = request.json
    expiry = d.get('expiry') or None          # '' → None → SQL NULL
    execute(
        "UPDATE reg_links SET title=%s,code=%s,url=%s,active=%s,max_uses=%s,used_count=%s,expiry=%s WHERE id=%s",
        (d['title'], d['code'], d.get('url',''), d.get('active',True), d.get('maxUses',0), d.get('usedCount',0), expiry, id)
    )
    audit('EDIT_REG_LINK', f"id={id} title={d['title']}")
    return jsonify({'ok': True})

@app.route('/api/reg-links/<int:id>', methods=['DELETE'])
def delete_reg_link(id):
    execute("DELETE FROM reg_links WHERE id=%s", (id,))
    audit('DELETE_REG_LINK', f"id={id}")
    return jsonify({'ok': True})

# ═══════════════════════════════════════════
# PUBLIC REGISTRATION API
# ═══════════════════════════════════════════
@app.route('/register')
@app.route('/register/<code>')
def register_page(code=None):
    return send_from_directory(os.path.join(ROOT, 'html'), 'register.html')

@app.route('/api/register/<code>')
def validate_reg_link(code):
    link = query("SELECT * FROM reg_links WHERE code=%s AND active=true", (code,), one=True)
    if not link:
        return jsonify({'ok': False, 'error': 'This registration link is invalid or inactive.'}), 404
    if link['expiry'] and link['expiry'] < _dt.date.today():
        return jsonify({'ok': False, 'error': 'This registration link has expired.'}), 410
    if link['max_uses'] and link['used_count'] >= link['max_uses']:
        return jsonify({'ok': False, 'error': 'This registration link has reached its maximum uses.'}), 410
    return jsonify({'ok': True, 'title': link['title'], 'code': link['code']})

@app.route('/api/register/<code>', methods=['POST'])
def submit_registration(code):
    link = query("SELECT * FROM reg_links WHERE code=%s AND active=true", (code,), one=True)
    if not link:
        return jsonify({'ok': False, 'error': 'Invalid or inactive link.'}), 404
    if link['expiry'] and link['expiry'] < _dt.date.today():
        return jsonify({'ok': False, 'error': 'Link has expired.'}), 410
    if link['max_uses'] and link['used_count'] >= link['max_uses']:
        return jsonify({'ok': False, 'error': 'Link has reached maximum uses.'}), 410

    d = request.json or {}
    name = (d.get('name') or '').strip()
    if not name:
        return jsonify({'ok': False, 'error': 'Name is required.'}), 400

    n = lambda k: d.get(k) or None
    execute("""
        INSERT INTO pending_members (
            reg_link_code, name, member_type, uid, date_of_initiation, date_of_registration_jigyasu,
            date_of_first_initiation, date_of_second_initiation,
            date_of_birth, blood_group, caste, nationality, profession, ashram,
            mobile1, mobile2, landline, office_phone, email1, email2,
            address_line1, address_line2, address_line3, city, pincode, state, country,
            qualification, occupation, designation, organization,
            father_title, father_first_name, father_middle_name, father_last_name,
            father_branch, father_bslno, father_uid, father_doi, father_phone, father_city, father_state,
            mother_title, mother_first_name, mother_middle_name, mother_last_name,
            mother_branch, mother_bslno, mother_uid, mother_doi, mother_phone, mother_city, mother_state,
            spouse_title, spouse_first_name, spouse_middle_name, spouse_last_name,
            spouse_branch, spouse_bslno, spouse_uid, spouse_doi, spouse_phone, spouse_city, spouse_state,
            nee_first_name, nee_middle_name, nee_last_name,
            mahila_association_member, youth_member, associate_youth_member,
            junior_pre_initiate_member, senior_pre_initiate_member,
            crc_member, cca_member, sant_su_member,
            ref1_name, ref1_address, ref1_email, ref1_phone, ref1_branch, ref1_relation,
            ref2_name, ref2_address, ref2_email, ref2_phone, ref2_branch, ref2_relation,
            notes, seva_interests
        ) VALUES (
            %s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,
            %s,%s,%s,%s,%s,%s,
            %s,%s,%s,%s,%s,%s,%s,
            %s,%s,%s,%s,
            %s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,
            %s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,
            %s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,
            %s,%s,%s,
            %s,%s,%s,%s,%s,%s,%s,%s,
            %s,%s,%s,%s,%s,%s,
            %s,%s,%s,%s,%s,%s,
            %s,%s
        )
    """, (
        code, name, n('memberType'), n('uid'), n('dateOfInitiation'), n('dateOfRegistrationJigyasu'),
        n('dateOfFirstInitiation'), n('dateOfSecondInitiation'),
        n('dateOfBirth'), n('bloodGroup'), n('caste'), n('nationality'), n('profession'), n('ashram'),
        n('mobile1'), n('mobile2'), n('landline'), n('officePhone'), n('email1'), n('email2'),
        n('addressLine1'), n('addressLine2'), n('addressLine3'), n('city'), n('pincode'), n('state'), n('country'),
        n('qualification'), n('occupation'), n('designation'), n('organization'),
        n('fatherTitle'), n('fatherFirstName'), n('fatherMiddleName'), n('fatherLastName'),
        n('fatherBranch'), n('fatherBslno'), n('fatherUid'), n('fatherDoi'), n('fatherPhone'), n('fatherCity'), n('fatherState'),
        n('motherTitle'), n('motherFirstName'), n('motherMiddleName'), n('motherLastName'),
        n('motherBranch'), n('motherBslno'), n('motherUid'), n('motherDoi'), n('motherPhone'), n('motherCity'), n('motherState'),
        n('spouseTitle'), n('spouseFirstName'), n('spouseMiddleName'), n('spouseLastName'),
        n('spouseBranch'), n('spouseBslno'), n('spouseUid'), n('spouseDoi'), n('spousePhone'), n('spouseCity'), n('spouseState'),
        n('neeFirstName'), n('neeMiddleName'), n('neeLastName'),
        n('mahilaAssociationMember'), n('youthMember'), n('associateYouthMember'),
        n('juniorPreInitiateMember'), n('seniorPreInitiateMember'),
        n('crcMember'), n('ccaMember'), n('santSuMember'),
        n('ref1Name'), n('ref1Address'), n('ref1Email'), n('ref1Phone'), n('ref1Branch'), n('ref1Relation'),
        n('ref2Name'), n('ref2Address'), n('ref2Email'), n('ref2Phone'), n('ref2Branch'), n('ref2Relation'),
        n('notes'), n('sevaInterests')
    ))
    execute("UPDATE reg_links SET used_count=used_count+1 WHERE code=%s", (code,))
    audit('REGISTRATION_SUBMIT', f"code={code} name={name}")
    return jsonify({'ok': True, 'message': 'Registration submitted successfully! An administrator will review your details.'}), 201

@app.route('/api/pending-members')
def get_pending_members():
    return jsonify(query("SELECT * FROM pending_members ORDER BY submitted_at DESC"))

@app.route('/api/pending-members/<int:id>/approve', methods=['POST'])
def approve_pending_member(id):
    row = query("SELECT * FROM pending_members WHERE id=%s", (id,), one=True)
    if not row:
        return jsonify({'ok': False, 'error': 'Not found'}), 404

    actor = request.headers.get('X-User', 'unknown')
    
    uid = (row.get('uid') or '').strip() or f"PENDING-{id}"

    # Auto-assign BSL and category
    category = determine_member_category(row)
    bsl = get_next_bsl(category)
    
    # Check UID uniqueness based on member type
    if category == 'Superhumane':
        if not uid.startswith('PENDING-') and query("SELECT 1 FROM superhumane_details WHERE uid=%s", (uid,)):
            return jsonify({'ok': False, 'error': f"UID '{uid}' is already in use in Superhumane records"}), 409
        
        # Build address from parts
        addr_parts = [row.get('address_line1'), row.get('address_line2'), row.get('address_line3'),
                      row.get('city'), row.get('state'), row.get('pincode'), row.get('country')]
        full_address = ', '.join(p for p in addr_parts if p) or None
        
        # Build father name from parts
        father_name = ' '.join(filter(None, [row.get('father_title'), row.get('father_first_name'), 
                                              row.get('father_middle_name'), row.get('father_last_name')])) or None
        mother_name = ' '.join(filter(None, [row.get('mother_title'), row.get('mother_first_name'), 
                                              row.get('mother_middle_name'), row.get('mother_last_name')])) or None
        
        execute("""
            INSERT INTO superhumane_details (
                uid, bsl, name, date_of_birth, gender,
                father_name, father_contact, father_uid, father_doi,
                mother_name, mother_contact, mother_uid, mother_doi,
                address, member_type, branch
            ) VALUES (
                %s, %s, %s, %s, %s,
                %s, %s, %s, %s,
                %s, %s, %s, %s,
                %s, %s, %s
            )
        """, (
            uid, bsl, row['name'], row['date_of_birth'], None,
            father_name, row.get('father_phone'), row.get('father_uid'), row.get('father_doi'),
            mother_name, row.get('mother_phone'), row.get('mother_uid'), row.get('mother_doi'),
            full_address, 'Superhumane', 'Soaminagar Branch Delhi'
        ))
    else:
        # Initiated or Jigyasu → member_details
        if not uid.startswith('PENDING-') and query("SELECT 1 FROM member_details WHERE uid=%s", (uid,)):
            return jsonify({'ok': False, 'error': f"UID '{uid}' is already in use"}), 409
        
        execute("""
            INSERT INTO member_details (
                uid, name, date_of_initiation, date_of_registration_jigyasu,
                date_of_first_initiation, date_of_second_initiation,
                date_of_birth, blood_group, caste, nationality, profession, ashram,
                mobile1, mobile2, landline, office_phone, email1, email2,
                address_line1, address_line2, address_line3, city, pincode, state, country,
                qualification, occupation, designation, organization,
                father_title, father_first_name, father_middle_name, father_last_name,
                father_branch, father_bslno, father_uid, father_doi, father_phone, father_city, father_state,
                mother_title, mother_first_name, mother_middle_name, mother_last_name,
                mother_branch, mother_bslno, mother_uid, mother_doi, mother_phone, mother_city, mother_state,
                spouse_title, spouse_first_name, spouse_middle_name, spouse_last_name,
                spouse_branch, spouse_bslno, spouse_uid, spouse_doi, spouse_phone, spouse_city, spouse_state,
                nee_first_name, nee_middle_name, nee_last_name,
                mahila_association_member, youth_member, associate_youth_member,
                junior_pre_initiate_member, senior_pre_initiate_member,
                crc_member, cca_member, sant_su_member,
                ref1_name, ref1_address, ref1_email, ref1_phone, ref1_branch, ref1_relation,
                ref2_name, ref2_address, ref2_email, ref2_phone, ref2_branch, ref2_relation,
                record_status, bsl, category
            ) VALUES (
                %s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,
                %s,%s,%s,%s,%s,%s,
                %s,%s,%s,%s,%s,%s,%s,
                %s,%s,%s,%s,
                %s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,
                %s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,
                %s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,
                %s,%s,%s,
                %s,%s,%s,%s,%s,%s,%s,%s,
                %s,%s,%s,%s,%s,%s,
                %s,%s,%s,%s,%s,%s,
                'Activated', %s, %s
            )
        """, (
            uid, row['name'], row.get('date_of_initiation'), row.get('date_of_registration_jigyasu'),
            row.get('date_of_first_initiation'), row.get('date_of_second_initiation'),
            row['date_of_birth'], row['blood_group'], row['caste'], row['nationality'], row['profession'], row['ashram'],
            row['mobile1'], row['mobile2'], row['landline'], row['office_phone'], row['email1'], row['email2'],
            row['address_line1'], row['address_line2'], row['address_line3'], row['city'], row['pincode'], row['state'], row['country'],
            row['qualification'], row['occupation'], row['designation'], row['organization'],
            row['father_title'], row['father_first_name'], row['father_middle_name'], row['father_last_name'],
            row['father_branch'], row['father_bslno'], row['father_uid'], row['father_doi'], row['father_phone'], row['father_city'], row['father_state'],
            row['mother_title'], row['mother_first_name'], row['mother_middle_name'], row['mother_last_name'],
            row['mother_branch'], row['mother_bslno'], row['mother_uid'], row['mother_doi'], row['mother_phone'], row['mother_city'], row['mother_state'],
            row['spouse_title'], row['spouse_first_name'], row['spouse_middle_name'], row['spouse_last_name'],
            row['spouse_branch'], row['spouse_bslno'], row['spouse_uid'], row['spouse_doi'], row['spouse_phone'], row['spouse_city'], row['spouse_state'],
            row['nee_first_name'], row['nee_middle_name'], row['nee_last_name'],
            row['mahila_association_member'], row['youth_member'], row['associate_youth_member'],
            row['junior_pre_initiate_member'], row['senior_pre_initiate_member'],
            row['crc_member'], row['cca_member'], row['sant_su_member'],
            row['ref1_name'], row['ref1_address'], row['ref1_email'], row['ref1_phone'], row['ref1_branch'], row['ref1_relation'],
            row['ref2_name'], row['ref2_address'], row['ref2_email'], row['ref2_phone'], row['ref2_branch'], row['ref2_relation'],
            bsl, category
        ))
    
    execute(
        "UPDATE pending_members SET status='approved', reviewed_at=NOW(), reviewed_by=%s WHERE id=%s",
        (actor, id)
    )
    audit('APPROVE_REGISTRATION', f"id={id} name={row['name']} uid={uid or 'pending'} bsl={bsl} category={category}")

    # Send approval email if applicant provided an email
    to_email = row.get('email1') or row.get('email2')
    if to_email:
        def _r(field): return row.get(field) or '—'
        def _date(field):
            v = row.get(field)
            return str(v)[:10] if v else '—'
        def _row(label, value):
            if not value or value == '—': return ''
            return f'<tr><td style="padding:6px 12px;color:#666;font-size:0.82rem;white-space:nowrap;vertical-align:top">{label}</td><td style="padding:6px 12px;color:#222;font-size:0.82rem">{value}</td></tr>'

        addr_parts = [row.get('address_line1'), row.get('address_line2'), row.get('address_line3'),
                      row.get('city'), row.get('state'), row.get('pincode'), row.get('country')]
        address = ', '.join(p for p in addr_parts if p)

        father_name = ' '.join(filter(None, [row.get('father_title'), row.get('father_first_name'), row.get('father_middle_name'), row.get('father_last_name')]))
        mother_name = ' '.join(filter(None, [row.get('mother_title'), row.get('mother_first_name'), row.get('mother_middle_name'), row.get('mother_last_name')]))
        spouse_name = ' '.join(filter(None, [row.get('spouse_title'), row.get('spouse_first_name'), row.get('spouse_middle_name'), row.get('spouse_last_name')]))

        send_email(
            to_addr  = to_email,
            subject  = 'Your Registration is Approved — Soaminagar Branch Delhi',
            html_body= f"""
<!DOCTYPE html>
<html>
<body style="margin:0;padding:0;background:#f4f4f4;font-family:'Segoe UI',Arial,sans-serif">
<table width="100%" cellpadding="0" cellspacing="0">
<tr><td align="center" style="padding:40px 20px">
  <table width="600" style="background:#fff;border-radius:12px;overflow:hidden;box-shadow:0 4px 20px rgba(0,0,0,0.1)">
    <!-- Header -->
    <tr><td style="background:#1a2d5a;padding:32px;text-align:center">
      <div style="font-size:2rem">🙏</div>
      <h1 style="color:#f4a124;font-size:1.4rem;margin:10px 0 4px;font-family:Georgia,serif">Soaminagar Branch Delhi</h1>
      <p style="color:rgba(255,255,255,0.7);margin:0;font-size:0.85rem">Ra Dha Sva Aa Mi</p>
    </td></tr>
    <!-- Body -->
    <tr><td style="padding:32px 36px">
      <h2 style="color:#1a2d5a;margin:0 0 8px;font-size:1.2rem">Registration Approved ✅</h2>
      <p style="color:#444;line-height:1.6;margin:0 0 24px">
        Dear <strong>{_r('name')}</strong>, your membership registration has been
        <strong style="color:#16a34a">approved</strong>. Your details are below for your records.
      </p>

      <!-- Personal Details -->
      <p style="margin:0 0 8px;font-weight:600;color:#1a2d5a;font-size:0.85rem;text-transform:uppercase;letter-spacing:.04em">Personal Details</p>
      <table style="width:100%;border-collapse:collapse;margin-bottom:20px;background:#fafafa;border-radius:8px">
        {_row('Member UID', uid)}
        {_row('Branch Serial No.', bsl)}
        {_row('Member Type', category)}
        {_row('Full Name', _r('name'))}
        {_row('Date of Birth', _date('date_of_birth'))}
        {_row('Blood Group', _r('blood_group'))}
        {_row('Caste', _r('caste'))}
        {_row('Nationality', _r('nationality'))}
        {_row('Profession', _r('profession'))}
        {_row('Ashram', _r('ashram'))}
        {_row('Date of Initiation', _date('date_of_initiation'))}
        {_row('Date of Reg. (Jigyasu)', _date('date_of_registration_jigyasu'))}
        {_row('Date of 1st Initiation', _date('date_of_first_initiation'))}
        {_row('Date of 2nd Initiation', _date('date_of_second_initiation'))}
      </table>

      <!-- Contact -->
      <p style="margin:0 0 8px;font-weight:600;color:#1a2d5a;font-size:0.85rem;text-transform:uppercase;letter-spacing:.04em">Contact</p>
      <table style="width:100%;border-collapse:collapse;margin-bottom:20px;background:#fafafa;border-radius:8px">
        {_row('Mobile 1', _r('mobile1'))}
        {_row('Mobile 2', _r('mobile2'))}
        {_row('Landline', _r('landline'))}
        {_row('Email 1', _r('email1'))}
        {_row('Email 2', _r('email2'))}
        {_row('Address', address or '—')}
      </table>

      <!-- Professional -->
      <p style="margin:0 0 8px;font-weight:600;color:#1a2d5a;font-size:0.85rem;text-transform:uppercase;letter-spacing:.04em">Professional</p>
      <table style="width:100%;border-collapse:collapse;margin-bottom:20px;background:#fafafa;border-radius:8px">
        {_row('Qualification', _r('qualification'))}
        {_row('Occupation', _r('occupation'))}
        {_row('Designation', _r('designation'))}
        {_row('Organization', _r('organization'))}
      </table>

      <!-- Family -->
      {'<p style="margin:0 0 8px;font-weight:600;color:#1a2d5a;font-size:0.85rem;text-transform:uppercase;letter-spacing:.04em">Family</p><table style="width:100%;border-collapse:collapse;margin-bottom:20px;background:#fafafa;border-radius:8px">' + _row("Father", father_name or '—') + _row("Father UID", _r('father_uid')) + _row("Mother", mother_name or '—') + _row("Mother UID", _r('mother_uid')) + _row("Spouse", spouse_name or '—') + _row("Spouse UID", _r('spouse_uid')) + '</table>' if any([father_name, mother_name, spouse_name]) else ''}

      <p style="color:#888;font-size:0.82rem;margin:0">
        Please keep your Member UID safe. If any details are incorrect, contact your branch secretary.
      </p>

      <!-- Seva Interests -->
      {'<p style="margin:24px 0 8px;font-weight:600;color:#1a2d5a;font-size:0.85rem;text-transform:uppercase;letter-spacing:.04em">Seva Interests</p><div style="background:#fff8f0;border-left:4px solid #e07b29;padding:12px 16px;border-radius:6px;color:#555;font-size:0.85rem">' + _r('seva_interests') + '</div>' if row.get('seva_interests') else ''}
    </td></tr>
    <!-- Footer -->
    <tr><td style="background:#f8f8f8;padding:16px 36px;border-top:1px solid #eee;text-align:center">
      <p style="color:#aaa;font-size:0.76rem;margin:0">Automated message from Soaminagar Branch Delhi portal.</p>
    </td></tr>
  </table>
</td></tr>
</table>
</body>
</html>"""
        )

    return jsonify({'ok': True, 'uid': uid})

@app.route('/api/pending-members/<int:id>/reject', methods=['POST'])
def reject_pending_member(id):
    actor = request.headers.get('X-User', 'unknown')
    execute(
        "UPDATE pending_members SET status='rejected', reviewed_at=NOW(), reviewed_by=%s WHERE id=%s",
        (actor, id)
    )
    audit('REJECT_REGISTRATION', f"id={id}")
    return jsonify({'ok': True})

# ═══════════════════════════════════════════
# ANNOUNCEMENTS API
# ═══════════════════════════════════════════
@app.route('/api/announcements')
def get_announcements():
    return jsonify(query("SELECT * FROM announcements ORDER BY date DESC"))

@app.route('/api/announcements', methods=['POST'])
def add_announcement():
    d = request.json
    row = execute(
        "INSERT INTO announcements (title,content,date,author,priority,active) VALUES (%s,%s,%s,%s,%s,%s) RETURNING *",
        (d['title'], d.get('content',''), d.get('date'), d.get('author','Admin'), d.get('priority','medium'), d.get('active',True)),
        returning=True
    )
    audit('ADD_ANNOUNCEMENT', f"title={d['title']}")
    return jsonify(row), 201

@app.route('/api/announcements/<int:id>', methods=['PUT'])
def update_announcement(id):
    d = request.json
    execute(
        "UPDATE announcements SET title=%s,content=%s,date=%s,author=%s,priority=%s,active=%s WHERE id=%s",
        (d['title'], d.get('content',''), d.get('date'), d.get('author','Admin'), d.get('priority','medium'), d.get('active',True), id)
    )
    audit('EDIT_ANNOUNCEMENT', f"id={id} title={d['title']} active={d.get('active',True)}")
    return jsonify({'ok': True})

@app.route('/api/announcements/<int:id>', methods=['DELETE'])
def delete_announcement(id):
    execute("DELETE FROM announcements WHERE id=%s", (id,))
    audit('DELETE_ANNOUNCEMENT', f"id={id}")
    return jsonify({'ok': True})

# ═══════════════════════════════════════════
# ATTENDANCE APIs
# ═══════════════════════════════════════════
@app.route('/api/attendance/esatsang')
def get_esatsang():
    uid      = request.args.get('uid', '').strip()
    page     = request.args.get('page')
    per_page = request.args.get('per_page')

    if uid:
        # Member view: only their own records
        rows = query(
            "SELECT * FROM esatsang_attendance WHERE member_uid = %s ORDER BY attendance_date DESC NULLS LAST, id DESC",
            (uid,)
        )
        return jsonify({'rows': rows, 'total': len(rows), 'page': 1, 'pages': 1})

    if page is not None and per_page is not None:
        # Paginated fetch
        page     = max(1, int(page))
        per_page = max(1, int(per_page))
        offset   = (page - 1) * per_page
        total    = query("SELECT COUNT(*) AS n FROM esatsang_attendance", one=True)['n']
        rows     = query(
            "SELECT * FROM esatsang_attendance ORDER BY attendance_date DESC NULLS LAST, id DESC LIMIT %s OFFSET %s",
            (per_page, offset)
        )
        return jsonify({'rows': rows, 'total': total, 'page': page, 'pages': -(-total // per_page)})

    # No pagination params — return all records
    rows = query("SELECT * FROM esatsang_attendance ORDER BY attendance_date DESC NULLS LAST, id DESC")
    return jsonify({'rows': rows, 'total': len(rows), 'page': 1, 'pages': 1})

@app.route('/api/attendance/esatsang/upload', methods=['POST'])
def upload_esatsang_attendance():
    """Accept xlsx or csv, parse, and import into esatsang_attendance."""
    if 'file' not in request.files:
        return jsonify({'ok': False, 'error': 'No file provided.'}), 400
    f = request.files['file']
    filename = f.filename.lower()
    try:
        if filename.endswith('.csv'):
            import csv, io
            text = f.read().decode('utf-8-sig')
            reader = csv.DictReader(io.StringIO(text))
            raw = list(reader)
            headers = list(reader.fieldnames or [])
        elif filename.endswith(('.xlsx', '.xlsm', '.xls')):
            import openpyxl, io
            wb = openpyxl.load_workbook(io.BytesIO(f.read()), data_only=True)
            ws = wb.active
            all_rows = list(ws.iter_rows(values_only=True))
            if not all_rows:
                return jsonify({'ok': False, 'error': 'Empty workbook.'}), 400
            headers = [str(h).strip() if h is not None else '' for h in all_rows[0]]
            raw = [dict(zip(headers, [str(c).strip() if c is not None else '' for c in row])) for row in all_rows[1:]]
        else:
            return jsonify({'ok': False, 'error': 'Unsupported file type. Use .xlsx or .csv'}), 400

        def find_col(headers, *keywords):
            norm = lambda s: s.upper().replace(' ', '_').replace('-', '_')
            for h in headers:
                hn = norm(h)
                if any(kw in hn for kw in keywords):
                    return h
            return None

        col_date   = find_col(headers, 'ATTENDANCE_DATE', 'DATE', 'ATT_DATE')
        col_mid    = find_col(headers, 'MEMBER_ID', 'MID', 'MEMB_ID', 'MEMBER_I')
        col_event  = find_col(headers, 'EVENT_NAME', 'EVENT')
        col_first  = find_col(headers, 'FIRST_NAME', 'FIRST')
        col_mid_nm = find_col(headers, 'MIDDLE_NAME', 'MIDDLE')
        col_last   = find_col(headers, 'LAST_NAME', 'LAST')
        col_uid    = find_col(headers, 'MEMBER_UID', 'UID')
        col_branch = find_col(headers, 'BRANCH_NAME', 'BRANCH')
        col_loc    = find_col(headers, 'LOCATION', 'LOC')
        col_type   = find_col(headers, 'ATTENDANCE_TYPE', 'ATT_TYPE', 'TYPE')

        def sv(row, col):
            return str(row.get(col, '') or '').strip() if col else ''

        parsed = [
            {
                'date':      sv(r, col_date),
                'memberId':  sv(r, col_mid),
                'eventName': sv(r, col_event),
                'firstName': sv(r, col_first),
                'middleName':sv(r, col_mid_nm),
                'lastName':  sv(r, col_last),
                'memberUid': sv(r, col_uid),
                'branchName':sv(r, col_branch),
                'location':  sv(r, col_loc),
                'type':      sv(r, col_type),
            }
            for r in raw
            if sv(r, col_mid) or sv(r, col_first) or sv(r, col_uid)
        ]
        if not parsed:
            return jsonify({'ok': False, 'error': 'No valid data rows found.'}), 400

        # Bulk upsert — one DB round trip for all rows.
        # ON CONFLICT DO NOTHING skips exact duplicates (date+member_id+event).
        # Requires a unique index on those three columns (created if missing).
        from db import get_conn
        import psycopg2.extras, psycopg2
        conn = get_conn()
        try:
            with conn.cursor() as cur:
                # Ensure unique index exists for dedup
                cur.execute("""
                    CREATE UNIQUE INDEX IF NOT EXISTS uq_esatsang_date_member_event
                    ON esatsang_attendance (attendance_date, member_id, event_name)
                """)
                values = [
                    (r['date'] or None, r['memberId'], r['eventName'],
                     r['firstName'], r['middleName'], r['lastName'],
                     r['memberUid'], r['branchName'], r['location'], r['type'])
                    for r in parsed
                ]
                before = cur.execute("SELECT count(*) FROM esatsang_attendance") or 0
                cur.execute("SELECT count(*) FROM esatsang_attendance")
                before = cur.fetchone()[0]
                psycopg2.extras.execute_values(
                    cur,
                    """INSERT INTO esatsang_attendance
                       (attendance_date, member_id, event_name, first_name, middle_name,
                        last_name, member_uid, branch_name, location, attendance_type)
                       VALUES %s
                       ON CONFLICT (attendance_date, member_id, event_name) DO NOTHING""",
                    values,
                    page_size=2000
                )
                cur.execute("SELECT count(*) FROM esatsang_attendance")
                after = cur.fetchone()[0]
            conn.commit()
            inserted = after - before
            audit('UPLOAD_ESATSANG_ATTENDANCE', f"file={f.filename} inserted={inserted} skipped={len(parsed)-inserted}")
            return jsonify({'ok': True, 'count': inserted, 'skipped': len(parsed) - inserted}), 201
        except Exception as e:
            conn.rollback()
            raise
        finally:
            conn.close()
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500


@app.route('/api/attendance/branch')
def get_branch_attendance():
    return jsonify(query("SELECT * FROM branch_attendance ORDER BY member_name"))

@app.route('/api/attendance/branch/upload', methods=['POST'])
def upload_branch_attendance():
    """Accept an xlsx or csv file upload, parse it, and import into branch_attendance."""
    if 'file' not in request.files:
        return jsonify({'ok': False, 'error': 'No file provided.'}), 400

    f = request.files['file']
    filename = f.filename.lower()

    rows = []
    try:
        if filename.endswith('.csv'):
            import csv, io
            text = f.read().decode('utf-8-sig')  # strips BOM if present
            reader = csv.DictReader(io.StringIO(text))
            raw = list(reader)
            headers = reader.fieldnames or []
        elif filename.endswith(('.xlsx', '.xlsm', '.xls')):
            import openpyxl, io
            wb = openpyxl.load_workbook(io.BytesIO(f.read()), data_only=True)
            ws = wb.active
            all_rows = list(ws.iter_rows(values_only=True))
            if not all_rows:
                return jsonify({'ok': False, 'error': 'Empty workbook.'}), 400
            headers = [str(h).strip() if h is not None else '' for h in all_rows[0]]
            raw = [dict(zip(headers, [str(c).strip() if c is not None else '' for c in row])) for row in all_rows[1:]]
        else:
            return jsonify({'ok': False, 'error': 'Unsupported file type. Use .xlsx or .csv'}), 400

        # Flexible column detection (case-insensitive, underscore/space-insensitive)
        def find_col(headers, *keywords):
            norm = lambda s: s.upper().replace(' ','_').replace('-','_')
            for h in headers:
                hn = norm(h)
                if any(kw in hn for kw in keywords):
                    return h
            return None

        col_id   = find_col(headers, 'MEMBER_ID', 'MID', 'MEMB_ID')
        col_name = find_col(headers, 'MEMBER_NAME', 'NAME', 'FULL_NAME')
        col_att  = find_col(headers, 'EVENTS_ATT', 'ATTENDED', 'ATT')
        col_tot  = find_col(headers, 'TOTAL', 'TOT_EVENTS', 'MAX_EVENTS')
        col_br   = find_col(headers, 'BRANCH')

        missing = [k for k, v in [('Member ID', col_id), ('Member Name', col_name),
                                   ('Events Attended', col_att), ('Total Events', col_tot),
                                   ('Branch', col_br)] if not v]
        if missing:
            return jsonify({
                'ok': False,
                'error': f"Could not detect columns: {', '.join(missing)}. "
                         f"Columns found: {', '.join(headers)}"
            }), 400

        def safe_int(v):
            try: return int(float(str(v).strip() or 0))
            except: return 0

        parsed = [
            {
                'memberId':     str(r.get(col_id, '') or '').strip(),
                'memberName':   str(r.get(col_name, '') or '').strip(),
                'eventsAttended': safe_int(r.get(col_att, 0)),
                'totalEvents':  safe_int(r.get(col_tot, 0)),
                'branch':       str(r.get(col_br, '') or '').strip(),
            }
            for r in raw if str(r.get(col_name, '') or '').strip()
        ]

        if not parsed:
            return jsonify({'ok': False, 'error': 'No valid data rows found.'}), 400

        execute("DELETE FROM branch_attendance")
        for r in parsed:
            execute(
                "INSERT INTO branch_attendance (member_id, member_name, events_attended, total_branch_events, branch_name) VALUES (%s,%s,%s,%s,%s)",
                (r['memberId'], r['memberName'], r['eventsAttended'], r['totalEvents'], r['branch'])
            )
        audit('UPLOAD_BRANCH_ATTENDANCE', f"file={f.filename} rows={len(parsed)}")
        return jsonify({'ok': True, 'count': len(parsed)}), 201

    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500

@app.route('/api/attendance/haazri')
def get_haazri():
    return jsonify(query("SELECT * FROM haazri_attendance ORDER BY id"))

@app.route('/api/attendance/haazri', methods=['POST'])
def add_haazri_batch():
    """Accept an array of haazri records."""
    rows = request.json or []
    for r in rows:
        execute(
            "INSERT INTO haazri_attendance (uid,name,date_time_str,haazri_id,event_name,branch_name,geolocation_name) VALUES (%s,%s,%s,%s,%s,%s,%s)",
            (r.get('uid',''), r.get('name',''), r.get('dateTimeStr',''), r.get('haazriId',''), r.get('eventName',''), r.get('branchName',''), r.get('geolocationName',''))
        )
    return jsonify({'ok': True, 'count': len(rows)}), 201

# ═══════════════════════════════════════════
# DASHBOARD STATS (computed)
# ═══════════════════════════════════════════
@app.route('/api/dashboard/stats')
def dashboard_stats():
    stats = {}
    # Members — from member_details
    memberCount = query("SELECT count(*) as c FROM member_details", one=True)['c']
    superhumaneCount = query("SELECT count(*) as c FROM superhumane_details", one=True)['c']
    stats['memberCount']     = memberCount
    stats['superhumaneCount'] = superhumaneCount
    stats['totalMembers']    = memberCount + superhumaneCount  # Combined total
    stats['activeMembers']   = query("SELECT count(*) as c FROM member_details WHERE record_status='Activated'", one=True)['c']
    stats['transferIn']      = query("SELECT count(*) as c FROM member_details WHERE date_transfer_in IS NOT NULL", one=True)['c']
    stats['transferOut']     = query("SELECT count(*) as c FROM member_details WHERE date_transfer_out IS NOT NULL", one=True)['c']
    stats['expired']         = query("SELECT count(*) as c FROM member_details WHERE date_of_expire IS NOT NULL AND date_of_expire < CURRENT_DATE", one=True)['c']
    stats['pendingApprovals'] = query("SELECT count(*) as c FROM pending_members WHERE status='pending'", one=True)['c']
    # Reg links
    stats['activeRegLinks']  = query("SELECT count(*) as c FROM reg_links WHERE active=true", one=True)['c']
    # Zones
    stats['activeZones']     = query("SELECT count(*) as c FROM zones WHERE active=true", one=True)['c']
    stats['inactiveZones']   = query("SELECT count(*) as c FROM zones WHERE active=false", one=True)['c']
    stats['membersWithZone'] = query("SELECT count(*) as c FROM member_details WHERE city IS NOT NULL AND city != ''", one=True)['c']
    # Branch (no branches table, derive from member_details)
    stats['activeBranchCodes']   = query("SELECT count(DISTINCT branch_id_card_received) as c FROM member_details WHERE branch_id_card_received IS NOT NULL AND branch_id_card_received != ''", one=True)['c']
    stats['inactiveBranchCodes'] = 0
    stats['membersWithBranch']   = query("SELECT count(*) as c FROM member_details WHERE branch_id_card_received IS NOT NULL AND branch_id_card_received != ''", one=True)['c']
    return jsonify(stats)

@app.route('/api/dashboard/attendance-stats')
def dashboard_attendance_stats():
    esatsang_count = query("SELECT count(*) as c FROM esatsang_attendance", one=True)['c']
    audio = query("SELECT count(*) as c FROM esatsang_attendance WHERE attendance_type='AUDIO'", one=True)['c']
    video = query("SELECT count(*) as c FROM esatsang_attendance WHERE attendance_type='VIDEO'", one=True)['c']
    branch_total = query("SELECT count(*) as c FROM branch_attendance", one=True)['c']
    branch_attended = query("SELECT count(*) as c FROM branch_attendance WHERE events_attended > 0", one=True)['c']
    latest_date = query("SELECT MAX(attendance_date) as d FROM esatsang_attendance", one=True)
    return jsonify({
        'esatsangCount': esatsang_count,
        'audioCount': audio,
        'videoCount': video,
        'branchTotal': branch_total,
        'branchAttended': branch_attended,
        'latestDate': str(latest_date['d']) if latest_date and latest_date['d'] else ''
    })

# ═══════════════════════════════════════════
# CHANGE PASSWORD (authenticated user)
# ═══════════════════════════════════════════
@app.route('/api/auth/change-password', methods=['POST'])
def change_password():
    data = request.json or {}
    username = (data.get('username') or '').strip()
    old_password = data.get('oldPassword') or ''
    new_password = data.get('newPassword') or ''

    if len(new_password) < 6:
        return jsonify({'ok': False, 'error': 'New password must be at least 6 characters.'}), 400

    user = query("SELECT id FROM users WHERE username=%s AND password=%s", (username, old_password), one=True)
    if not user:
        return jsonify({'ok': False, 'error': 'Current password is incorrect.'}), 403

    execute("UPDATE users SET password=%s WHERE username=%s", (new_password, username))
    return jsonify({'ok': True})


# ═══════════════════════════════════════════
# SUPERHUMANE — children visible to parents
# ═══════════════════════════════════════════

@app.route('/api/my-children')
def my_children():
    """Return superhumane_details rows where father_uid or mother_uid matches the caller's UID."""
    uid = (request.args.get('uid') or '').strip()
    if not uid:
        return jsonify({'ok': False, 'error': 'uid is required'}), 400

    rows = query(
        """
        SELECT *
        FROM superhumane_details
        WHERE father_uid = %s OR mother_uid = %s
        ORDER BY name
        """,
        (uid, uid)
    )
    children = []
    for r in (rows or []):
        d = dict(r)
        for k, v in d.items():
            if isinstance(v, (_dt.date, _dt.datetime)):
                d[k] = v.isoformat()
        children.append(d)
    return jsonify({'ok': True, 'children': children})


@app.route('/api/all-superhumane')
def all_superhumane():
    """Return all superhumane_details rows (admin view)."""
    rows = query("SELECT * FROM superhumane_details ORDER BY sno")
    children = []
    for r in (rows or []):
        d = dict(r)
        for k, v in d.items():
            if isinstance(v, (_dt.date, _dt.datetime)):
                d[k] = v.isoformat()
        children.append(d)
    return jsonify({'ok': True, 'children': children})


@app.route('/api/superhumane', methods=['POST'])
def add_superhumane():
    import psycopg2
    d = request.json or {}
    uid  = (d.get('uid')  or '').strip()
    name = (d.get('name') or '').strip()
    if not uid:  return jsonify({'ok': False, 'error': 'UID is required.'}), 400
    if not name: return jsonify({'ok': False, 'error': 'Name is required.'}), 400
    or_none = lambda k: d.get(k) or None
    try:
        execute("""
            INSERT INTO superhumane_details
              (uid, name, member_type, gender, bsl, phase, branch, date_of_birth,
               form_check, uid_check, address, comments,
               father_name, father_uid, father_contact, father_doi,
               mother_name, mother_uid, mother_contact, mother_doi,
               grandfather_name, grandfather_uid, grandfather_contact,
               grandmother_name, grandmother_uid, grandmother_contact,
               date_entry_scheme, date_exit_scheme)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s, %s,%s,%s,%s, %s,%s,%s,%s, %s,%s,%s,%s, %s,%s,%s, %s,%s,%s, %s,%s)
        """, (
            uid, name, d.get('member_type'), d.get('gender'), d.get('bsl'),
            d.get('phase'), d.get('branch'), or_none('date_of_birth'),
            d.get('form_check'), d.get('uid_check'), d.get('address'), d.get('comments'),
            d.get('father_name'), d.get('father_uid'), d.get('father_contact'), or_none('father_doi'),
            d.get('mother_name'), d.get('mother_uid'), d.get('mother_contact'), or_none('mother_doi'),
            d.get('grandfather_name'), d.get('grandfather_uid'), d.get('grandfather_contact'),
            d.get('grandmother_name'), d.get('grandmother_uid'), d.get('grandmother_contact'),
            or_none('date_entry_scheme'), or_none('date_exit_scheme')
        ))
        audit('ADD_SUPERHUMANE', f"uid={uid} name={name}")
        return jsonify({'ok': True, 'uid': uid}), 201
    except psycopg2.errors.UniqueViolation:
        return jsonify({'ok': False, 'error': f"A Superhumane record with UID '{uid}' already exists."}), 409
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500


@app.route('/api/superhumane/<uid>', methods=['PUT'])
def update_superhumane(uid):
    d = request.json or {}
    or_none = lambda k: d.get(k) or None

    # Status-only toggle: { exit: true } → set exit date; { exit: false } → clear it
    if set(d.keys()) <= {'exit'}:
        if d.get('exit'):
            execute("UPDATE superhumane_details SET date_exit_scheme=%s WHERE uid=%s",
                    (_dt.date.today().isoformat(), uid))
            audit('EXIT_SUPERHUMANE_SCHEME', f"uid={uid}")
        else:
            execute("UPDATE superhumane_details SET date_exit_scheme=NULL WHERE uid=%s", (uid,))
            audit('REACTIVATE_SUPERHUMANE', f"uid={uid}")
        return jsonify({'ok': True})

    # Full field update
    execute("""
        UPDATE superhumane_details SET
          name=%s, member_type=%s, gender=%s, bsl=%s, phase=%s, branch=%s,
          date_of_birth=%s, form_check=%s, uid_check=%s, address=%s, comments=%s,
          father_name=%s, father_uid=%s, father_contact=%s, father_doi=%s,
          mother_name=%s, mother_uid=%s, mother_contact=%s, mother_doi=%s,
          grandfather_name=%s, grandfather_uid=%s, grandfather_contact=%s,
          grandmother_name=%s, grandmother_uid=%s, grandmother_contact=%s,
          date_entry_scheme=%s, date_exit_scheme=%s
        WHERE uid=%s
    """, (
        d.get('name'), d.get('member_type'), d.get('gender'), d.get('bsl'),
        d.get('phase'), d.get('branch'), or_none('date_of_birth'),
        d.get('form_check'), d.get('uid_check'), d.get('address'), d.get('comments'),
        d.get('father_name'), d.get('father_uid'), d.get('father_contact'), or_none('father_doi'),
        d.get('mother_name'), d.get('mother_uid'), d.get('mother_contact'), or_none('mother_doi'),
        d.get('grandfather_name'), d.get('grandfather_uid'), d.get('grandfather_contact'),
        d.get('grandmother_name'), d.get('grandmother_uid'), d.get('grandmother_contact'),
        or_none('date_entry_scheme'), or_none('date_exit_scheme'),
        uid
    ))
    audit('EDIT_SUPERHUMANE', f"uid={uid} name={d.get('name','')}")
    return jsonify({'ok': True})


@app.route('/api/superhumane/upload', methods=['POST'])
def upload_superhumane():
    """Accept xlsx or csv, parse Superhumane sheet columns, and upsert into superhumane_details."""
    if 'file' not in request.files:
        return jsonify({'ok': False, 'error': 'No file provided.'}), 400
    f = request.files['file']
    filename = f.filename.lower()
    
    try:
        all_raw = []
        headers = []
        
        if filename.endswith('.csv'):
            import csv, io
            text = f.read().decode('utf-8-sig')
            reader = csv.DictReader(io.StringIO(text))
            all_raw = list(reader)
            headers = list(reader.fieldnames or [])
        elif filename.endswith(('.xlsx', '.xlsm', '.xls')):
            import openpyxl, io
            wb = openpyxl.load_workbook(io.BytesIO(f.read()), data_only=True)
            
            # Try to find Superhumane sheet, or use active sheet
            ws = None
            for sheet_name in wb.sheetnames:
                if 'superhumane' in sheet_name.lower() or 'sant' in sheet_name.lower():
                    ws = wb[sheet_name]
                    break
            if ws is None:
                ws = wb.active
            
            all_rows = list(ws.iter_rows(values_only=True))
            if not all_rows:
                return jsonify({'ok': False, 'error': 'Empty workbook.'}), 400
            
            # Find header row (look for S.No., UID, Name pattern)
            header_row_idx = None
            for i, row in enumerate(all_rows[:10]):
                if row and any(cell is not None and str(cell).strip() for cell in row):
                    row_str = ' '.join(str(c or '').upper() for c in row)
                    if ('UID' in row_str or 'S.NO' in row_str or 'SNO' in row_str) and 'NAME' in row_str:
                        header_row_idx = i
                        break
            
            if header_row_idx is None:
                return jsonify({'ok': False, 'error': 'Could not find header row with UID/Name columns.'}), 400
            
            raw_headers = [str(h).strip() if h is not None else '' for h in all_rows[header_row_idx]]
            # Make headers unique by appending position for duplicates
            seen = {}
            headers = []
            for idx, h in enumerate(raw_headers):
                if h in seen:
                    seen[h] += 1
                    headers.append(f"{h}___{idx}")  # unique key: e.g. "UID___15"
                else:
                    seen[h] = 1
                    headers.append(h)
            all_raw = [dict(zip(headers, [c for c in row])) for row in all_rows[header_row_idx + 1:]]
        else:
            return jsonify({'ok': False, 'error': 'Unsupported file type. Use .xlsx or .csv'}), 400

        # Column mapping for Superhumane sheet
        # Excel headers → DB columns (handle duplicate column names like UID, Contact Number)
        def normalize(h):
            # Strip the unique suffix we added (e.g., "UID___15" -> "UID")
            base = h.split('___')[0] if '___' in h else h
            return base.upper().replace('\xa0', ' ').replace('\n', ' ').strip()
        
        # Build column mapping by position (since there are duplicate column names)
        col_map = {}
        uid_count = 0
        contact_count = 0
        doi_count = 0
        
        for idx, h in enumerate(headers):
            norm = normalize(h)
            if 'S.NO' in norm or 'SNO' in norm:
                col_map[h] = 'sno'
            elif norm == 'MEMBER TYPE':
                col_map[h] = 'member_type'
            elif norm == 'NAME' and 'FATHER' not in norm and 'MOTHER' not in norm and 'GRAND' not in norm:
                col_map[h] = 'name'
            elif 'FORM CHECK' in norm:
                col_map[h] = 'form_check'
            elif 'UID CHECK' in norm:
                col_map[h] = 'uid_check'
            elif norm == 'BSL':
                col_map[h] = 'bsl'
            elif norm in ('M/F', 'GENDER'):
                col_map[h] = 'gender'
            elif norm == 'COMMENTS':
                col_map[h] = 'comments'
            elif norm == 'UID':
                uid_count += 1
                if uid_count == 1:
                    col_map[h] = 'uid'
                elif uid_count == 2:
                    col_map[h] = 'father_uid'
                elif uid_count == 3:
                    col_map[h] = 'mother_uid'
            elif 'DOB' in norm or 'DATE OF BIRTH' in norm:
                col_map[h] = 'date_of_birth'
            elif norm == 'AGE':
                pass  # Skip age - computed field
            elif norm == 'PHASE':
                col_map[h] = 'phase'
            elif norm == 'BRANCH':
                col_map[h] = 'branch'
            # ── GRANDPARENT checks MUST come before PARENT checks ──
            elif 'GRAND' in norm and 'FATHER' in norm:
                if 'UID' in norm:
                    col_map[h] = 'grandfather_uid'
                elif 'CONTACT' in norm:
                    col_map[h] = 'grandfather_contact'
                else:
                    col_map[h] = 'grandfather_name'
            elif 'GRAND' in norm and 'MOTHER' in norm:
                if 'UID' in norm:
                    col_map[h] = 'grandmother_uid'
                elif 'CONTACT' in norm:
                    col_map[h] = 'grandmother_contact'
                else:
                    col_map[h] = 'grandmother_name'
            # ── PARENT checks ──
            elif 'FATHER' in norm and 'NAME' in norm:
                col_map[h] = 'father_name'
            elif 'MOTHER' in norm and 'NAME' in norm:
                col_map[h] = 'mother_name'
            elif 'CONTACT' in norm:
                contact_count += 1
                if contact_count == 1:
                    col_map[h] = 'father_contact'
                elif contact_count == 2:
                    col_map[h] = 'mother_contact'
                # Note: grandparent contacts handled above
            elif 'DOI' in norm or 'DOR' in norm:
                doi_count += 1
                if doi_count == 1:
                    col_map[h] = 'father_doi'
                elif doi_count == 2:
                    col_map[h] = 'mother_doi'
            elif norm == 'ADDRESS':
                col_map[h] = 'address'
            elif 'ENTRY' in norm and 'SCHEME' in norm:
                col_map[h] = 'date_entry_scheme'
            elif 'EXIT' in norm and 'SCHEME' in norm:
                col_map[h] = 'date_exit_scheme'

        DATE_COLS = {'date_of_birth', 'father_doi', 'mother_doi', 'date_entry_scheme', 'date_exit_scheme'}

        def parse_date(val):
            if val is None:
                return None
            if isinstance(val, (_dt.date, _dt.datetime)):
                return val if isinstance(val, _dt.date) else val.date()
            s = str(val).strip()
            if not s or s.lower() in ('none', 'null', 'nat', ''):
                return None
            for fmt in ('%d-%m-%Y', '%d/%m/%Y', '%Y-%m-%d', '%m/%d/%Y', '%d-%b-%Y'):
                try:
                    return _dt.datetime.strptime(s, fmt).date()
                except ValueError:
                    pass
            try:
                serial = float(s)
                if 1 < serial < 100000:
                    return (_dt.datetime(1899, 12, 30) + _dt.timedelta(days=int(serial))).date()
            except:
                pass
            return None

        def clean_val(val, db_col):
            if val is None:
                return None
            if db_col == 'sno' or db_col == 'phase':
                try:
                    return int(float(val))
                except:
                    return None
            if db_col in DATE_COLS:
                return parse_date(val)
            if db_col == 'gender':
                v = str(val).strip().upper()
                return 'M' if v in ('M', 'MALE') else 'F' if v in ('F', 'FEMALE') else v[:1] if v else None
            s = str(val).replace('\xa0', ' ').strip()
            return s if s and s.lower() not in ('none', 'null') else None

        # Process rows
        parsed_rows = []
        for row in all_raw:
            row_data = {}
            uid_val = None
            for excel_h, db_col in col_map.items():
                val = row.get(excel_h)
                cleaned = clean_val(val, db_col)
                if db_col == 'uid':
                    uid_val = cleaned
                row_data[db_col] = cleaned
            
            if uid_val:
                parsed_rows.append(row_data)

        if not parsed_rows:
            return jsonify({'ok': False, 'error': 'No valid data rows with UID found.'}), 400

        # Upsert into superhumane_details
        from db import get_conn
        conn = get_conn()
        try:
            with conn.cursor() as cur:
                cur.execute("SELECT count(*) FROM superhumane_details")
                before = cur.fetchone()[0]

                for row_data in parsed_rows:
                    cols = list(row_data.keys())
                    vals = [row_data[c] for c in cols]
                    set_clause = ', '.join(f"{c}=EXCLUDED.{c}" for c in cols if c != 'uid')
                    sql = f"""
                        INSERT INTO superhumane_details ({', '.join(cols)})
                        VALUES ({', '.join(['%s'] * len(cols))})
                        ON CONFLICT (uid) DO UPDATE SET {set_clause}
                    """
                    cur.execute(sql, vals)

                cur.execute("SELECT count(*) FROM superhumane_details")
                after = cur.fetchone()[0]
            conn.commit()
            inserted = after - before
            updated = len(parsed_rows) - inserted
            audit('UPLOAD_SUPERHUMANE', f"file={f.filename} new={inserted} updated={updated} total={len(parsed_rows)}")
            return jsonify({
                'ok': True,
                'count': len(parsed_rows),
                'inserted': inserted,
                'updated': updated,
                'columns_found': list(col_map.values())
            }), 201
        except Exception as e:
            conn.rollback()
            raise
        finally:
            conn.close()
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'ok': False, 'error': str(e)}), 500


# ═══════════════════════════════════════════
# MEMBERS UPLOAD (Excel/CSV)
# ═══════════════════════════════════════════
@app.route('/api/members/upload', methods=['POST'])
def upload_members():
    """Accept xlsx or csv, parse all columns dynamically, and upsert into member_details.
    For Excel files with multiple sheets, use ?sheets=FormA,Jigyasus to select specific sheets.
    If no sheets param, processes all sheets that have member-like data (UID column).
    """
    if 'file' not in request.files:
        return jsonify({'ok': False, 'error': 'No file provided.'}), 400
    f = request.files['file']
    filename = f.filename.lower()
    selected_sheets = request.args.get('sheets', '').split(',') if request.args.get('sheets') else None
    
    try:
        all_raw = []  # Combined rows from all sheets
        all_headers = set()
        sheets_processed = []
        
        if filename.endswith('.csv'):
            import csv, io
            text = f.read().decode('utf-8-sig')
            reader = csv.DictReader(io.StringIO(text))
            all_raw = list(reader)
            all_headers = set(reader.fieldnames or [])
            sheets_processed = ['CSV']
        elif filename.endswith(('.xlsx', '.xlsm', '.xls')):
            import openpyxl, io
            file_bytes = f.read()
            wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
            
            # Determine which sheets to process
            sheets_to_process = []
            for sheet_name in wb.sheetnames:
                # Skip sheets that are clearly not member data
                lower_name = sheet_name.lower()
                if lower_name in ('data', 'lookup', 'reference', 'sheet1', 'superhumane', 'sant-su', 'santsu'):
                    continue
                # If user specified sheets, only process those
                if selected_sheets and sheet_name not in selected_sheets:
                    continue
                sheets_to_process.append(sheet_name)
            
            # If no sheets selected and none auto-detected, try all sheets
            if not sheets_to_process:
                sheets_to_process = wb.sheetnames
            
            for sheet_name in sheets_to_process:
                ws = wb[sheet_name]
                all_rows = list(ws.iter_rows(values_only=True))
                if not all_rows:
                    continue
                
                # Find header row (first row with UID/Name columns)
                header_row_idx = None
                for i, row in enumerate(all_rows[:10]):  # Check first 10 rows
                    if row and any(cell is not None and str(cell).strip() for cell in row):
                        row_str = ' '.join(str(c or '').upper() for c in row)
                        if 'UID' in row_str and ('NAME' in row_str or 'SL' in row_str):
                            header_row_idx = i
                            break
                
                if header_row_idx is None:
                    continue  # Skip sheets without proper headers
                
                headers = [str(h).strip() if h is not None else '' for h in all_rows[header_row_idx]]
                all_headers.update(headers)
                
                # Add rows from this sheet
                for row in all_rows[header_row_idx + 1:]:
                    row_dict = dict(zip(headers, [c for c in row]))
                    # Add sheet name as a reference
                    row_dict['_source_sheet'] = sheet_name
                    all_raw.append(row_dict)
                
                sheets_processed.append(f"{sheet_name} ({len(all_rows) - header_row_idx - 1} rows)")
            
            if not sheets_processed:
                return jsonify({'ok': False, 'error': 'No valid sheets found with member data (need UID column).'}), 400
        else:
            return jsonify({'ok': False, 'error': 'Unsupported file type. Use .xlsx or .csv'}), 400
        
        headers = list(all_headers)

        # Column name mapping: Excel header variations → DB column name
        COLUMN_MAP = {
            'SL': 'sl',
            'UID': 'uid',
            'NAME': 'name',
            'DATE OF INITIATION': 'date_of_initiation',
            'BSL': 'bsl',
            'DATE OF BIRTH': 'date_of_birth',
            'DATE OF REGISTRATION (JIGYASU)': 'date_of_registration_jigyasu',
            'DATE OF REGISTRATION JIGYASU': 'date_of_registration_jigyasu',
            'DATE OF FIRST INITIATION': 'date_of_first_initiation',
            'DATE OF SECOND INITIATION': 'date_of_second_initiation',
            'CASTE': 'caste',
            'NATIONALITY': 'nationality',
            'QUALIFICATION': 'qualification',
            'OCCUPATION': 'occupation',
            'DESIGNATION': 'designation',
            'ORGANIZATION': 'organization',
            'PROFESSION': 'profession',
            'ADDRESS LINE1': 'address_line1',
            'ADDRESS LINE 1': 'address_line1',
            'ADDRESS LINE2': 'address_line2',
            'ADDRESS LINE 2': 'address_line2',
            'ADDRESS LINE3': 'address_line3',
            'ADDRESS LINE 3': 'address_line3',
            'CITY': 'city',
            'PINCODE': 'pincode',
            'STATE': 'state',
            'COUNTRY': 'country',
            'SN/EXT': 'sn_ext',
            'SN_EXT': 'sn_ext',
            'ASHRAM': 'ashram',
            'EMAIL-1': 'email1',
            'EMAIL1': 'email1',
            'EMAIL 1': 'email1',
            'EMAIL-2': 'email2',
            'EMAIL2': 'email2',
            'EMAIL 2': 'email2',
            'MOBILE-1': 'mobile1',
            'MOBILE1': 'mobile1',
            'MOBILE 1': 'mobile1',
            'MOBILE-2': 'mobile2',
            'MOBILE2': 'mobile2',
            'MOBILE 2': 'mobile2',
            'LANDLINE': 'landline',
            'OFFICE PHONE': 'office_phone',
            'BLOOD GROUP': 'blood_group',
            'BRANCH I. CARD RECEIVED': 'branch_id_card_received',
            'BRANCH ID CARD RECEIVED': 'branch_id_card_received',
            'BRANCH I CARD RECEIVED': 'branch_id_card_received',
            'FATHER TITLE': 'father_title',
            'FATHER FIRST NAME': 'father_first_name',
            'FATHER MIDDLE NAME': 'father_middle_name',
            'FATHER LAST NAME': 'father_last_name',
            'FATHER BRANCH': 'father_branch',
            'FATHER BSLNO': 'father_bslno',
            'FATHER BSL': 'father_bslno',
            'FATHER UID': 'father_uid',
            'FATHER DOI': 'father_doi',
            'FATHER PHONE NO.': 'father_phone',
            'FATHER PHONE': 'father_phone',
            'FATHER CITY': 'father_city',
            'FATHER STATE': 'father_state',
            'MOTHER TITLE': 'mother_title',
            'MOTHER FIRST NAME': 'mother_first_name',
            'MOTHER MIDDLE NAME': 'mother_middle_name',
            'MOTHER LAST NAME': 'mother_last_name',
            'MOTHER BRANCH': 'mother_branch',
            'MOTHER BSLNO': 'mother_bslno',
            'MOTHER BSL': 'mother_bslno',
            'MOTHER UID': 'mother_uid',
            'MOTHER DOI': 'mother_doi',
            'MOTHER PHONE NO.': 'mother_phone',
            'MOTHER PHONE': 'mother_phone',
            'MOTHER CITY': 'mother_city',
            'MOTHER STATE': 'mother_state',
            'SPOUSE TITLE': 'spouse_title',
            'SPOUSE FIRST NAME': 'spouse_first_name',
            'SPOUSE MIDDLE NAME': 'spouse_middle_name',
            'SPOUSE LAST NAME': 'spouse_last_name',
            'SPOUSE BRANCH': 'spouse_branch',
            'SPOUSE BSLNO': 'spouse_bslno',
            'SPOUSE BSL': 'spouse_bslno',
            'SPOUSE UID': 'spouse_uid',
            'SPOUSE DOI': 'spouse_doi',
            'SPOUSE PHONE NO.': 'spouse_phone',
            'SPOUSE PHONE': 'spouse_phone',
            'SPOUSE CITY': 'spouse_city',
            'SPOUSE STATE': 'spouse_state',
            'MAHILA ASSOCIATION MEMBER': 'mahila_association_member',
            'YOUTH MEMBER': 'youth_member',
            'ASSOCIATE YOUTH MEMBER': 'associate_youth_member',
            'JUNIOR PRE INITIATE MEMBER': 'junior_pre_initiate_member',
            'SENIOR PRE INITIATE MEMBER': 'senior_pre_initiate_member',
            'CRC MEMBER': 'crc_member',
            'CCA MEMBER': 'cca_member',
            'SANT-SU MEMBER': 'sant_su_member',
            'SANT SU MEMBER': 'sant_su_member',
            'NEE (FIRST NAME)': 'nee_first_name',
            'NEE FIRST NAME': 'nee_first_name',
            'NEE (MIDDLE NAME)': 'nee_middle_name',
            'NEE MIDDLE NAME': 'nee_middle_name',
            'NEE (LAST NAME)': 'nee_last_name',
            'NEE LAST NAME': 'nee_last_name',
            'REF-1 (NAME)': 'ref1_name',
            'REF-1 NAME': 'ref1_name',
            'REF1 NAME': 'ref1_name',
            'REF-1 (ADDRESS)': 'ref1_address',
            'REF-1 ADDRESS': 'ref1_address',
            'REF1 ADDRESS': 'ref1_address',
            'REF-1 (E-MAIL)': 'ref1_email',
            'REF-1 EMAIL': 'ref1_email',
            'REF1 EMAIL': 'ref1_email',
            'REF-1 (MOBILE/PHONE)': 'ref1_phone',
            'REF-1 PHONE': 'ref1_phone',
            'REF1 PHONE': 'ref1_phone',
            'REF-1 (BRANCH NAME)': 'ref1_branch',
            'REF-1 BRANCH': 'ref1_branch',
            'REF1 BRANCH': 'ref1_branch',
            'REF-1 (RELATION)': 'ref1_relation',
            'REF-1 RELATION': 'ref1_relation',
            'REF1 RELATION': 'ref1_relation',
            'REF-2 (NAME)': 'ref2_name',
            'REF-2 NAME': 'ref2_name',
            'REF2 NAME': 'ref2_name',
            'REF-2 (ADDRESS)': 'ref2_address',
            'REF-2 ADDRESS': 'ref2_address',
            'REF2 ADDRESS': 'ref2_address',
            'REF-3 (E-MAIL)': 'ref2_email',
            'REF-2 (E-MAIL)': 'ref2_email',
            'REF-2 EMAIL': 'ref2_email',
            'REF2 EMAIL': 'ref2_email',
            'REF-4 (MOBILE/PHONE)': 'ref2_phone',
            'REF-2 (MOBILE/PHONE)': 'ref2_phone',
            'REF-2 PHONE': 'ref2_phone',
            'REF2 PHONE': 'ref2_phone',
            'REF-5 (BRANCH NAME)': 'ref2_branch',
            'REF-2 (BRANCH NAME)': 'ref2_branch',
            'REF-2 BRANCH': 'ref2_branch',
            'REF2 BRANCH': 'ref2_branch',
            'REF-6 (RELATION)': 'ref2_relation',
            'REF-2 (RELATION)': 'ref2_relation',
            'REF-2 RELATION': 'ref2_relation',
            'REF2 RELATION': 'ref2_relation',
            'DOR (YOUTH)': 'dor_youth',
            'DOR YOUTH': 'dor_youth',
            'DATE OF INITIATION (FOR NEW INITIATE)': 'date_of_initiation_new',
            'DATE OF INITIATION NEW': 'date_of_initiation_new',
            'DATE OF TRANSFER IN': 'date_transfer_in',
            'TRANSFER FROM (BRANCH)': 'transfer_from_branch',
            'TRANSFER FROM BRANCH': 'transfer_from_branch',
            'DATE OF TRANSFER OUT': 'date_transfer_out',
            'TRANSFER TO (BRANCH)': 'transfer_to_branch',
            'TRANSFER TO BRANCH': 'transfer_to_branch',
            'DATE OF EXPIRE': 'date_of_expire',
            'RECORD STATUS': 'record_status',
            'PROFESSION CODE': 'profession_code',
            'COMMUNICATION GRID CODE': 'communication_grid_code',
            'CATEGORY': 'category',
            'GENDER': 'gender',
            'MARITAL STATUS': 'marital_status',
            'PREVIOUS BRANCH': 'previous_branch',
        }

        DATE_COLUMNS = {
            'date_of_initiation', 'date_of_birth', 'date_of_registration_jigyasu',
            'date_of_first_initiation', 'date_of_second_initiation',
            'father_doi', 'mother_doi', 'spouse_doi',
            'dor_youth', 'date_of_initiation_new',
            'date_transfer_in', 'date_transfer_out', 'date_of_expire',
        }

        def normalize_header(h):
            return h.upper().replace('\xa0', ' ').replace('_', ' ').strip()

        def parse_date(val):
            if val is None:
                return None
            if isinstance(val, (_dt.date, _dt.datetime)):
                return val if isinstance(val, _dt.date) else val.date()
            s = str(val).strip()
            if not s or s.lower() in ('none', 'null', 'nat', ''):
                return None
            for fmt in ('%d-%m-%Y', '%d/%m/%Y', '%Y-%m-%d', '%m/%d/%Y', '%d-%b-%Y', '%d %b %Y'):
                try:
                    return _dt.datetime.strptime(s, fmt).date()
                except ValueError:
                    pass
            # Try parsing as Excel serial date
            try:
                serial = float(s)
                if 1 < serial < 100000:
                    return (_dt.datetime(1899, 12, 30) + _dt.timedelta(days=int(serial))).date()
            except (ValueError, TypeError):
                pass
            return None

        def clean_value(val, db_col):
            if val is None:
                return None
            if db_col == 'sl':
                try:
                    return int(float(val))
                except (TypeError, ValueError):
                    return None
            if db_col in DATE_COLUMNS:
                return parse_date(val)
            s = str(val).replace('\xa0', ' ').strip() if not isinstance(val, str) else val.replace('\xa0', ' ').strip()
            return s if s and s.lower() not in ('none', 'null') else None

        # Map Excel headers to DB columns
        header_to_db = {}
        for h in headers:
            norm = normalize_header(h)
            if norm in COLUMN_MAP:
                header_to_db[h] = COLUMN_MAP[norm]

        if 'uid' not in header_to_db.values():
            # Try to find UID column more loosely
            for h in headers:
                if 'UID' in h.upper():
                    header_to_db[h] = 'uid'
                    break

        if 'uid' not in header_to_db.values():
            return jsonify({'ok': False, 'error': 'Could not find UID column in the file. Headers found: ' + ', '.join(headers[:20])}), 400

        # Process rows from all sheets
        parsed_rows = []
        for row in all_raw:
            row_data = {}
            uid_val = None
            for excel_h, db_col in header_to_db.items():
                val = row.get(excel_h)
                cleaned = clean_value(val, db_col)
                if db_col == 'uid':
                    uid_val = cleaned
                row_data[db_col] = cleaned
            
            # Skip empty rows (no UID)
            if uid_val:
                # Set default record_status if not provided
                if 'record_status' not in row_data or not row_data['record_status']:
                    row_data['record_status'] = 'Activated'
                parsed_rows.append(row_data)

        if not parsed_rows:
            return jsonify({'ok': False, 'error': 'No valid data rows with UID found.'}), 400

        # Upsert into member_details
        from db import get_conn
        import psycopg2.extras
        conn = get_conn()
        try:
            with conn.cursor() as cur:
                cur.execute("SELECT count(*) FROM member_details")
                before = cur.fetchone()[0]

                for row_data in parsed_rows:
                    cols = list(row_data.keys())
                    vals = [row_data[c] for c in cols]
                    # Upsert: INSERT ... ON CONFLICT (uid) DO UPDATE
                    set_clause = ', '.join(f"{c}=EXCLUDED.{c}" for c in cols if c != 'uid')
                    sql = f"""
                        INSERT INTO member_details ({', '.join(cols)})
                        VALUES ({', '.join(['%s'] * len(cols))})
                        ON CONFLICT (uid) DO UPDATE SET {set_clause}
                    """
                    cur.execute(sql, vals)

                cur.execute("SELECT count(*) FROM member_details")
                after = cur.fetchone()[0]
            conn.commit()
            inserted = after - before
            updated = len(parsed_rows) - inserted
            audit('UPLOAD_MEMBERS', f"file={f.filename} sheets={sheets_processed} new={inserted} updated={updated} total={len(parsed_rows)}")
            return jsonify({
                'ok': True,
                'count': len(parsed_rows),
                'inserted': inserted,
                'updated': updated,
                'sheets_processed': sheets_processed,
                'columns_found': list(header_to_db.values())
            }), 201
        except Exception as e:
            conn.rollback()
            raise
        finally:
            conn.close()
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'ok': False, 'error': str(e)}), 500


if __name__ == '__main__':
    # Run schema + seed on first start
    from db import get_conn
    import seed as seed_module

    print("Setting up database...")
    conn = get_conn()
    cur = conn.cursor()
    with open(os.path.join(os.path.dirname(__file__), 'schema.sql')) as f:
        sql = f.read()
    # psycopg2 execute() does not support multiple statements; run each individually
    for statement in sql.split(';'):
        stmt = statement.strip()
        if stmt:
            cur.execute(stmt)
    conn.commit()
    cur.close()
    conn.close()

    print("Seeding data...")
    seed_module.seed()

    print("Starting server on http://localhost:5001")
    app.run(host='0.0.0.0', port=5001, debug=True)
